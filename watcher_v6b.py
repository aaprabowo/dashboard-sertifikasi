"""
=============================================================
  Dashboard Monitoring Terpadu — Excel Watcher Script
  File: basis data.xlsx + sheet: output (targets)
=============================================================
"""

import json, os, sys, time, logging
from datetime import datetime
from collections import defaultdict, OrderedDict

try:
    import openpyxl
except ImportError:
    print("[ERROR] pip install openpyxl"); sys.exit(1)

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    print("[ERROR] pip install watchdog"); sys.exit(1)


# =============================================================
#  KONFIGURASI
# =============================================================

EXCEL_FILE  = r"C:\Users\NamaAnda\OneDrive - Organisasi\Data Sertifikasi\basis data.xlsx"
OUTPUT_JSON = os.path.join(os.path.dirname(__file__), "data", "dashboard_data.json")
SHEET_MAP   = {
    "JFKN":     "JFKN",
    "SAK":      "SAK",
    "AC":       "AC",
    "Beasiswa": "Beasiswa",
    "USKP":     "USKP",
    "PBJ":      "PBJ",
    "PengTes":  "PengTes",
}
DEBOUNCE_SECONDS = 3
NILAI_LULUS      = {"lulus"}

# Mapping nama bulan Inggris -> angka
MONTH_EN = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
    "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,
    "sep":9,"oct":10,"nov":11,"dec":12,
}


# =============================================================
#  LOGGING
# =============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(os.path.dirname(__file__), "watcher.log"), encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)


# =============================================================
#  HELPERS
# =============================================================

def col(headers, *candidates):
    for c in candidates:
        t = c.lower().strip()
        for i, h in enumerate(headers):
            if h and str(h).lower().strip() == t:
                return i
    return -1

def val(row, idx):
    if idx == -1 or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()

def is_lulus(v):
    return v.lower() in NILAI_LULUS

def count_dist(items):
    d = defaultdict(int)
    for x in items:
        if x: d[x] += 1
    return dict(sorted(d.items(), key=lambda x: x[1], reverse=True))

def to_rows(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows: return [], []
    headers = [str(c).lower().strip() if c else "" for c in all_rows[0]]
    data    = [r for r in all_rows[1:] if any(v is not None for v in r)]
    return headers, data

def empty(team, error=None):
    return {
        "team": team, "type": "count", "total": 0,
        "lulus": 0, "tidak_lulus": 0, "dalam_proses": 0, "pct_lulus": 0,
        "total_angkatan": 0, "total_hadir": 0, "months_by_year": {}, "detail": {}, "error": error,
    }

def extract_month_year(value):
    """Ekstrak (bulan, tahun) dari berbagai format. Return (None, None) jika gagal."""
    if value is None: return None, None
    # datetime/date object
    if hasattr(value, 'month'): return value.month, value.year
    s = str(value).strip()
    if not s or s.lower() == 'none': return None, None
    sl = s.lower()
    # Nama bulan Inggris saja (tanpa tahun) — tahun tidak diketahui
    if sl in MONTH_EN: return MONTH_EN[sl], None
    for name, num in MONTH_EN.items():
        if len(name) >= 3 and name in sl:
            # Coba ambil tahun dari string yang sama
            import re
            m = re.search(r'\b(20\d{2})\b', s)
            yr = int(m.group(1)) if m else None
            return num, yr
    # Format angka
    for sep in ['/', '-', '.']:
        parts = s.split(sep)
        if len(parts) == 3:
            try:
                nums = [int(p) for p in parts]
                if nums[0] > 31: return nums[1], nums[0]   # yyyy-mm-dd
                if nums[2] > 31: return nums[1], nums[2]   # dd/mm/yyyy
            except: pass
    return None, None

def add_month_year(months_dict, value, lulus=False):
    """
    Tambah satu baris ke struktur months_by_year.
    months_dict: { year: { month: {total, lulus} } }
    value: raw nilai dari kolom tanggal
    """
    m, y = extract_month_year(value)
    if m is None: return
    if y is None: y = datetime.now().year   # fallback tahun sekarang
    if y not in months_dict: months_dict[y] = {}
    if m not in months_dict[y]: months_dict[y][m] = {"total": 0, "lulus": 0}
    months_dict[y][m]["total"] += 1
    if lulus: months_dict[y][m]["lulus"] += 1

def sort_batches(batch_list):
    """Pertahankan urutan asli (insertion order dari dict)."""
    return batch_list

def wpct(d):
    return [{"label": k, "total": v["total"], "lulus": v["lulus"],
             "pct": round(v["lulus"]/v["total"]*100,1) if v["total"] > 0 else 0}
            for k, v in d.items()]


# =============================================================
#  BACA SHEET OUTPUT (TARGETS)
# =============================================================

def read_targets(wb):
    """Baca sheet 'output' untuk target peserta per tim per tahun."""
    targets = {}
    sheet_name = next((s for s in wb.sheetnames if s.lower() == 'output'), None)
    if not sheet_name:
        log.warning("Sheet 'output' tidak ditemukan — progress bar target tidak akan tampil.")
        return targets
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return targets
    headers = [str(c).strip() if c else "" for c in rows[0]]
    # Cari kolom 'ujian' dan kolom tahun (4 digit angka)
    col_ujian = next((i for i, h in enumerate(headers) if h.lower() == 'ujian'), -1)
    if col_ujian == -1:
        log.warning("Kolom 'ujian' tidak ditemukan di sheet output.")
        return targets
    year_cols = {h: i for i, h in enumerate(headers) if h.isdigit() and len(h) == 4}
    for row in rows[1:]:
        if not any(v is not None for v in row): continue
        team = str(row[col_ujian]).strip() if row[col_ujian] else ""
        if not team: continue
        targets[team] = {}
        for yr, idx in year_cols.items():
            if idx < len(row) and row[idx] is not None:
                try: targets[team][int(yr)] = int(row[idx])
                except: pass
    log.info(f"Targets dibaca: {targets}")
    return targets


# =============================================================
#  PROCESSORS PER TIM
# =============================================================

def process_JFKN(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("JFKN", "Sheet kosong")
    i_lulus   = col(headers, "kelulusan")
    i_ue1     = col(headers, "ue1")
    i_ue2     = col(headers, "ue2")
    i_jenjang = col(headers, "jenjang_target")
    i_jf      = col(headers, "jf_target")
    i_tgl     = col(headers, "tglukom", "tgl_ukom", "tglukom")
    if i_lulus == -1: return empty("JFKN", "Kolom 'Kelulusan' tidak ditemukan")

    total = lulus = dalam_proses = 0
    ue1_d = OrderedDict(); ue2_d = OrderedDict()
    jenjang_vals = []; jf_vals = []
    months_by_year = {}
    # Untuk filter — simpan per baris agar bisa filter di detail
    rows_detail = []

    for r in rows:
        raw_kel = val(r, i_lulus)
        # Skip baris template/contoh (misal: 'lulus/tidak lulus')
        if raw_kel and '/' in raw_kel:
            continue
        total += 1
        if not raw_kel:
            # Kolom kosong = belum ada data / dalam proses
            dalam_proses += 1
            vl = False
        else:
            vl = is_lulus(raw_kel)
            if vl: lulus += 1

        u1 = val(r, i_ue1)
        u2 = val(r, i_ue2)
        jn = val(r, i_jenjang)
        jf = val(r, i_jf)

        if u1:
            if u1 not in ue1_d: ue1_d[u1] = {"total":0,"lulus":0}
            ue1_d[u1]["total"] += 1
            if vl: ue1_d[u1]["lulus"] += 1
        if u2:
            if u2 not in ue2_d: ue2_d[u2] = {"total":0,"lulus":0}
            ue2_d[u2]["total"] += 1
            if vl: ue2_d[u2]["lulus"] += 1
        if jn: jenjang_vals.append(jn)
        if jf:  jf_vals.append(jf)

        rows_detail.append({"ue1": u1, "ue2": u2, "jenjang": jn, "jf": jf, "lulus": raw_kel})

        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        add_month_year(months_by_year, raw_tgl, lulus=vl)

    # Hitung hanya dari baris yang sudah ada hasil (bukan dalam_proses)
    ada_hasil = total - dalam_proses
    pct = round(lulus / ada_hasil * 100, 1) if ada_hasil > 0 else 0

    # Sort UE1/UE2 by total terbesar
    ue1_sorted = sorted(ue1_d.items(), key=lambda x: x[1]["total"], reverse=True)
    ue2_sorted = sorted(ue2_d.items(), key=lambda x: x[1]["total"], reverse=True)

    log.info(f"[JFKN] {total} peserta, {lulus} lulus ({pct}%), {dalam_proses} dalam proses")
    return {
        "team":"JFKN","type":"pct","total":total,"total_hadir":total,
        "lulus":lulus,"tidak_lulus":ada_hasil-lulus,"dalam_proses":dalam_proses,
        "pct_lulus":pct,"total_angkatan":0,
        "targets": targets.get("JFKN",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "ue1":     [{"label":k,"total":v["total"],"lulus":v["lulus"],
                         "pct":round(v["lulus"]/v["total"]*100,1) if v["total"]>0 else 0}
                        for k,v in ue1_sorted],
            "ue2":     [{"label":k,"total":v["total"],"lulus":v["lulus"],
                         "pct":round(v["lulus"]/v["total"]*100,1) if v["total"]>0 else 0}
                        for k,v in ue2_sorted],
            "jenjang": count_dist(jenjang_vals),
            "jf":      count_dist(jf_vals),
            "filter_options": {
                "jenjang": sorted(set(x for x in jenjang_vals if x)),
                "jf":      sorted(set(x for x in jf_vals if x)),
            },
            "rows": rows_detail,
        }
    }


def process_SAK(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("SAK", "Sheet kosong")
    i_lulus  = col(headers, "kelulusan")
    i_period = col(headers, "certification_period_name")
    i_mon    = col(headers, "certification_period_month")
    i_yr     = col(headers, "certification_period_year")
    i_lokasi = col(headers, "exam_location_name")
    i_reg    = col(headers, "registrasi")
    i_vdok   = col(headers, "verifikasi dokumen")
    i_vpay   = col(headers, "verifikasi pembayaran")
    i_hadir  = col(headers, "kehadiran")
    if i_lulus == -1: return empty("SAK", "Kolom 'kelulusan' tidak ditemukan")

    total = lulus = 0
    n_reg = n_vdok = n_vpay = n_hadir = 0
    batch_stat = OrderedDict()
    lokasi_vals = []
    months_by_year = {}

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl: lulus += 1
        # funnel
        if i_reg   != -1 and val(r, i_reg):   n_reg   += 1
        if i_vdok  != -1 and val(r, i_vdok):  n_vdok  += 1
        if i_vpay  != -1 and val(r, i_vpay):  n_vpay  += 1
        if i_hadir != -1 and val(r, i_hadir):
            h = val(r, i_hadir).lower()
            if "hadir" in h and "tidak" not in h: n_hadir += 1
        # batch
        p = val(r, i_period)
        if p:
            if p not in batch_stat: batch_stat[p] = {"total":0,"lulus":0}
            batch_stat[p]["total"] += 1
            if vl: batch_stat[p]["lulus"] += 1
        # lokasi
        lk = val(r, i_lokasi)
        if lk: lokasi_vals.append(lk)
        # month — dari certification_period_month (teks nama bulan)
        raw_mon = r[i_mon] if i_mon != -1 and i_mon < len(r) else None
        add_month_year(months_by_year, raw_mon, lulus=vl)

    pct   = round(lulus/total*100,1) if total > 0 else 0
    trend = [{"batch":k,"total":v["total"],"lulus":v["lulus"],
               "pct":round(v["lulus"]/v["total"]*100,1) if v["total"]>0 else 0}
             for k,v in batch_stat.items()]
    log.info(f"[SAK] {total} peserta, {n_hadir} hadir, {lulus} lulus ({pct}%)")
    return {
        "team":"SAK","type":"pct","total":total,"total_hadir":n_hadir,"lulus":lulus,
        "tidak_lulus":total-lulus,"pct_lulus":pct,
        "total_angkatan":len(batch_stat),
        "targets":targets.get("SAK",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "trend": trend,
            "lokasi": count_dist(lokasi_vals),
            "funnel":{
                "pendaftar": total,
                "verif_dokumen": n_vdok,
                "verif_pembayaran": n_vpay,
                "hadir": n_hadir,
                "lulus": lulus,
            }
        }
    }


def process_AC(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("AC", "Sheet kosong")
    i_hasil = col(headers, "hasil penilaian kompetensi")
    i_ue1   = col(headers, "ue1")
    i_batch = col(headers, "batch")
    i_tgl   = col(headers, "tanggal ac")

    hasil_vals = []; ue1_vals = []
    batch_stat = OrderedDict()
    months_by_year = {}
    # Data per tahun untuk filter popup
    batch_by_year = {}   # {year: {batch: {total,optimal,cukup,kurang}}}
    ue1_by_year   = {}   # {year: {ue1: count}}
    hasil_by_year = {}   # {year: {hasil: count}}

    for r in rows:
        h = val(r, i_hasil)
        if h: hasil_vals.append(h)
        u = val(r, i_ue1)
        if u: ue1_vals.append(u)
        b = val(r, i_batch)

        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        _, yr = extract_month_year(raw_tgl)
        if yr is None: yr = datetime.now().year

        if b:
            # Global batch stat
            if b not in batch_stat: batch_stat[b] = {"total":0,"optimal":0,"cukup":0,"kurang":0}
            batch_stat[b]["total"] += 1
            hl = h.lower() if h else ""
            if "optimal" in hl and "cukup" not in hl and "kurang" not in hl:
                batch_stat[b]["optimal"] += 1
            elif "cukup" in hl: batch_stat[b]["cukup"] += 1
            elif "kurang" in hl: batch_stat[b]["kurang"] += 1

            # Per-year batch stat
            if yr not in batch_by_year: batch_by_year[yr] = OrderedDict()
            if b not in batch_by_year[yr]: batch_by_year[yr][b] = {"total":0,"optimal":0,"cukup":0,"kurang":0}
            batch_by_year[yr][b]["total"] += 1
            if "optimal" in hl and "cukup" not in hl and "kurang" not in hl:
                batch_by_year[yr][b]["optimal"] += 1
            elif "cukup" in hl: batch_by_year[yr][b]["cukup"] += 1
            elif "kurang" in hl: batch_by_year[yr][b]["kurang"] += 1

        # Per-year UE1
        if u:
            if yr not in ue1_by_year: ue1_by_year[yr] = {}
            ue1_by_year[yr][u] = ue1_by_year[yr].get(u, 0) + 1

        # Per-year hasil
        if h:
            if yr not in hasil_by_year: hasil_by_year[yr] = {}
            hasil_by_year[yr][h] = hasil_by_year[yr].get(h, 0) + 1

        add_month_year(months_by_year, raw_tgl)

    total = len(rows)
    log.info(f"[AC] {total} peserta")
    return {
        "team":"AC","type":"count","total":total,"total_hadir":total,"lulus":0,"tidak_lulus":0,
        "dalam_proses":0,"pct_lulus":0,"total_angkatan":len(batch_stat),
        "targets":targets.get("AC",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "hasil": count_dist(hasil_vals),
            "ue1":   count_dist(ue1_vals),
            "batch_stat":    [{"batch":k,**v} for k,v in batch_stat.items()],
            "batch_by_year": {str(yr): [{"batch":k,**v} for k,v in bmap.items()]
                              for yr, bmap in batch_by_year.items()},
            "ue1_by_year":   {str(yr): d for yr, d in ue1_by_year.items()},
            "hasil_by_year": {str(yr): d for yr, d in hasil_by_year.items()},
        }
    }


def process_Beasiswa(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("Beasiswa", "Sheet kosong")
    i_status = col(headers, "status")
    i_prodi  = col(headers, "prodi")
    i_ue1    = col(headers, "ue1")
    i_univ   = col(headers, "universitas")
    i_negara = col(headers, "negara")
    i_jenj   = col(headers, "jenjang")
    i_bea    = col(headers, "beasiswa")
    i_batch  = col(headers, "batch")
    i_tgl    = col(headers, "tgl_regis", "tgl regis", "tanggal registrasi")

    status_v=[]; prodi_v=[]; ue1_v=[]; univ_v=[]; negara_v=[]; jenj_v=[]; bea_v=[]
    rows_detail=[]; batch_set=set()
    months_by_year = {}

    for r in rows:
        st=val(r,i_status); status_v.append(st) if st else None
        pr=val(r,i_prodi);  prodi_v.append(pr)  if pr else None
        u1=val(r,i_ue1);    ue1_v.append(u1)    if u1 else None
        un=val(r,i_univ);   univ_v.append(un)   if un else None
        ng=val(r,i_negara); negara_v.append(ng) if ng else None
        jn=val(r,i_jenj);   jenj_v.append(jn)   if jn else None
        bv=val(r,i_bea);    bea_v.append(bv)    if bv else None
        b=val(r,i_batch);
        if b: batch_set.add(b)
        rows_detail.append({"ue1":u1,"jenjang":jn,"negara":ng,"prodi":pr,"universitas":un,"beasiswa":bv,"status":st})
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        add_month_year(months_by_year, raw_tgl)

    total = len(rows)
    log.info(f"[Beasiswa] {total} penerima")
    return {
        "team":"Beasiswa","type":"count","total":total,"total_hadir":total,"lulus":0,"tidak_lulus":0,
        "pct_lulus":0,"total_angkatan":len(batch_set),
        "targets":targets.get("Beasiswa",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "status":count_dist(status_v),"prodi":count_dist(prodi_v),
            "ue1":count_dist(ue1_v),"universitas":count_dist(univ_v),
            "negara":count_dist(negara_v),"jenjang":count_dist(jenj_v),
            "beasiswa":count_dist(bea_v),
            "filter_options":{
                "ue1":    sorted(set(x for x in ue1_v    if x)),
                "jenjang":sorted(set(x for x in jenj_v   if x)),
                "negara": sorted(set(x for x in negara_v if x)),
            },
            "rows": rows_detail,
        }
    }


def process_USKP(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("USKP", "Sheet kosong")
    i_lulus  = col(headers, "kelulusan")
    i_batch  = col(headers, "batch")
    i_lokasi = col(headers, "lokasi")
    i_verif  = col(headers, "verifikasi")
    i_hadir  = col(headers, "kehadiran")
    i_tgl    = col(headers, "tgl_ujian")
    if i_lulus == -1: return empty("USKP", "Kolom 'kelulusan' tidak ditemukan")

    total = lulus = 0
    n_lolos_verif = n_hadir = 0
    batch_stat = OrderedDict()
    lokasi_vals = []
    months_by_year = {}

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl: lulus += 1
        # funnel
        if i_verif != -1:
            v = val(r, i_verif).lower()
            if "lolos" in v and "tidak" not in v: n_lolos_verif += 1
        if i_hadir != -1:
            h = val(r, i_hadir).lower()
            if "hadir" in h and "tidak" not in h: n_hadir += 1
        # batch
        b = val(r, i_batch)
        if b:
            if b not in batch_stat: batch_stat[b] = {"total":0,"lulus":0}
            batch_stat[b]["total"] += 1
            if vl: batch_stat[b]["lulus"] += 1
        lk = val(r, i_lokasi)
        if lk: lokasi_vals.append(lk)
        # month
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        add_month_year(months_by_year, raw_tgl, lulus=vl)

    pct   = round(lulus/total*100,1) if total > 0 else 0
    trend = [{"batch":k,"total":v["total"],"lulus":v["lulus"],
               "pct":round(v["lulus"]/v["total"]*100,1) if v["total"]>0 else 0}
             for k,v in batch_stat.items()]
    log.info(f"[USKP] {total} peserta, {n_hadir} hadir, {lulus} lulus ({pct}%)")
    return {
        "team":"USKP","type":"pct","total":total,"total_hadir":n_hadir,"lulus":lulus,
        "tidak_lulus":total-lulus,"pct_lulus":pct,
        "total_angkatan":len(batch_stat),
        "targets":targets.get("USKP",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "trend": trend,
            "lokasi": count_dist(lokasi_vals),
            "funnel":{
                "pendaftar": total,
                "lolos_verifikasi": n_lolos_verif,
                "hadir": n_hadir,
                "lulus": lulus,
            },
            "batch_list": sorted(batch_stat.keys()),
        }
    }


def process_PBJ(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("PBJ", "Sheet kosong")
    i_lulus = col(headers, "kelulusan")
    i_batch = col(headers, "batch")
    i_jenis = col(headers, "jenis")
    i_verif = col(headers, "verifikasi")
    i_hadir = col(headers, "kehadiran")
    i_tgl   = col(headers, "tgl_ujian")
    if i_lulus == -1: return empty("PBJ", "Kolom 'kelulusan' tidak ditemukan")

    total = lulus = 0
    n_lolos_verif = n_hadir = 0
    batch_stat = OrderedDict()
    jenis_vals = []
    months_by_year = {}

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl: lulus += 1
        if i_verif != -1:
            v = val(r, i_verif).lower()
            if "lolos" in v and "tidak" not in v: n_lolos_verif += 1
        if i_hadir != -1:
            h = val(r, i_hadir).lower()
            if "hadir" in h and "tidak" not in h: n_hadir += 1
        b = val(r, i_batch)
        if b:
            if b not in batch_stat: batch_stat[b] = {"total":0,"lulus":0}
            batch_stat[b]["total"] += 1
            if vl: batch_stat[b]["lulus"] += 1
        j = val(r, i_jenis)
        if j: jenis_vals.append(j)
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        add_month_year(months_by_year, raw_tgl, lulus=vl)

    pct   = round(lulus/total*100,1) if total > 0 else 0
    trend = [{"batch":k,"total":v["total"],"lulus":v["lulus"],
               "pct":round(v["lulus"]/v["total"]*100,1) if v["total"]>0 else 0}
             for k,v in batch_stat.items()]
    log.info(f"[PBJ] {total} peserta, {n_hadir} hadir, {lulus} lulus ({pct}%)")
    return {
        "team":"PBJ","type":"pct","total":total,"total_hadir":n_hadir,"lulus":lulus,
        "tidak_lulus":total-lulus,"pct_lulus":pct,
        "total_angkatan":len(batch_stat),
        "targets":targets.get("PBJ",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "trend": trend,
            "jenis": count_dist(jenis_vals),
            "funnel":{
                "pendaftar": total,
                "lolos_verifikasi": n_lolos_verif,
                "hadir": n_hadir,
                "lulus": lulus,
            },
            "batch_list": sorted(batch_stat.keys()),
        }
    }


def process_PengTes(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return empty("PengTes", "Sheet kosong")
    i_ujian = col(headers, "ujian")
    i_ue1   = col(headers, "ue1")
    i_batch = col(headers, "batch")
    i_tgl   = col(headers, "tgl_ujian")

    ujian_vals=[]; ue1_vals=[]
    batch_stat = OrderedDict()
    months_by_year = {}

    for r in rows:
        u = val(r, i_ujian)
        if u: ujian_vals.append(u)
        u1 = val(r, i_ue1)
        if u1: ue1_vals.append(u1)
        b = val(r, i_batch)
        if b:
            if b not in batch_stat: batch_stat[b] = 0
            batch_stat[b] += 1
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        add_month_year(months_by_year, raw_tgl)

    total = len(rows)
    log.info(f"[PengTes] {total} peserta")
    return {
        "team":"PengTes","type":"count","total":total,"total_hadir":total,"lulus":0,"tidak_lulus":0,
        "pct_lulus":0,"total_angkatan":len(batch_stat),
        "targets":targets.get("PengTes",{}),"error":None,
        "months_by_year": months_by_year,
        "detail":{
            "ujian": count_dist(ujian_vals),
            "ue1": count_dist(ue1_vals),
            "batch_stat":[{"batch":k,"total":v} for k,v in batch_stat.items()],
        }
    }


PROCESSORS = {
    "JFKN":     process_JFKN,
    "SAK":      process_SAK,
    "AC":       process_AC,
    "Beasiswa": process_Beasiswa,
    "USKP":     process_USKP,
    "PBJ":      process_PBJ,
    "PengTes":  process_PengTes,
}


# =============================================================
#  PROSES SEMUA SHEET
# =============================================================

def process_all_sheets():
    if not os.path.exists(EXCEL_FILE):
        log.error(f"File tidak ditemukan: {EXCEL_FILE}"); return
    try:
        log.info(f"Membuka: {os.path.basename(EXCEL_FILE)}")
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        log.info(f"Sheet tersedia: {wb.sheetnames}")
        targets = read_targets(wb)
        teams_data = []
        for team_name, sheet_name in SHEET_MAP.items():
            if sheet_name not in wb.sheetnames:
                log.warning(f"[{team_name}] Sheet '{sheet_name}' tidak ditemukan.")
                teams_data.append(empty(team_name, f"Sheet '{sheet_name}' tidak ditemukan"))
                continue
            proc = PROCESSORS.get(team_name)
            teams_data.append(proc(wb[sheet_name], targets) if proc else empty(team_name, "Processor tidak ditemukan"))
        wb.close()
    except Exception as e:
        log.error(f"Gagal membuka file Excel: {e}"); return

    pct_teams      = [t for t in teams_data if t["type"] == "pct"]
    total_peserta  = sum(t["total"]       for t in pct_teams)
    total_lulus    = sum(t["lulus"]       for t in pct_teams)
    total_hadir    = sum(t.get("total_hadir", t["total"]) for t in teams_data)
    avg_lulus      = round(total_lulus/total_peserta*100,1) if total_peserta > 0 else 0
    total_angkatan = sum(t.get("total_angkatan",0) for t in teams_data)

    # Capaian output tahunan: total_hadir semua tim / total target semua tim (tahun sekarang)
    cur_year = datetime.now().year
    total_target = sum(
        t.get("targets",{}).get(cur_year, 0) for t in teams_data
    )
    pct_capaian = round(total_hadir / total_target * 100, 1) if total_target > 0 else 0

    # Tim yang sudah mencapai ≥100% target
    tim_mencapai = sum(
        1 for t in teams_data
        if t.get("targets",{}).get(cur_year, 0) > 0 and
           t.get("total_hadir", t["total"]) >= t["targets"][cur_year]
    )

    # Beasiswa & AC untuk kartu bawah
    beasiswa_t = next((t for t in teams_data if t["team"]=="Beasiswa"), None)
    ac_t       = next((t for t in teams_data if t["team"]=="AC"), None)

    output = {
        "generated_at":    datetime.now().strftime("%d %B %Y, %H:%M:%S"),
        "generated_ts":    datetime.now().isoformat(),
        "total_peserta":   total_peserta,
        "total_lulus":     total_lulus,
        "avg_pct_lulus":   avg_lulus,
        "total_tim":       len(SHEET_MAP),
        "total_angkatan":  total_angkatan,
        "total_hadir":     total_hadir,
        "total_target":    total_target,
        "pct_capaian":     pct_capaian,
        "tim_mencapai":    tim_mencapai,
        "total_beasiswa":  beasiswa_t["total"] if beasiswa_t else 0,
        "total_ac":        ac_t["total"]       if ac_t       else 0,
        "teams":           teams_data,
    }
    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    log.info(f"JSON diperbarui -> {OUTPUT_JSON}")
    log.info(f"  {total_peserta} peserta | capaian {pct_capaian}% | {tim_mencapai}/{len(teams_data)} tim capai target")


# =============================================================
#  FILE WATCHER
# =============================================================

class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self): self._last = 0
    def on_modified(self, event): self._handle(event)
    def on_created(self, event):  self._handle(event)
    def _handle(self, event):
        if event.is_directory: return
        if os.path.normcase(os.path.abspath(event.src_path)) != \
           os.path.normcase(os.path.abspath(EXCEL_FILE)): return
        now = time.time()
        if now - self._last < DEBOUNCE_SECONDS: return
        self._last = now
        log.info("Perubahan terdeteksi -> memproses ulang...")
        time.sleep(1.5)
        process_all_sheets()


# =============================================================
#  MAIN
# =============================================================

def main():
    log.info("=" * 60)
    log.info("  Dashboard Monitoring Terpadu — Watcher Aktif")
    log.info("=" * 60)
    log.info(f"  File  : {EXCEL_FILE}")
    log.info(f"  Output: {OUTPUT_JSON}")
    log.info(f"  Sheet : {list(SHEET_MAP.values())}")
    log.info("  Ctrl+C untuk berhenti.")
    log.info("=" * 60)
    if not os.path.exists(EXCEL_FILE):
        log.error(f"File tidak ditemukan: {EXCEL_FILE}"); sys.exit(1)
    log.info("Memproses data awal...")
    process_all_sheets()
    watch_dir = os.path.dirname(os.path.abspath(EXCEL_FILE))
    handler = ExcelChangeHandler()
    observer = Observer()
    observer.schedule(handler, path=watch_dir, recursive=False)
    observer.start()
    log.info(f"Watcher aktif — memantau: {watch_dir}")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        log.info("Watcher dihentikan."); observer.stop()
    observer.join()

if __name__ == "__main__":
    main()
