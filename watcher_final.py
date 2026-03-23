"""
=============================================================
  Dashboard Monitoring Terpadu — Excel Watcher Script
  Cara A: OneDrive Sync + File Watcher
=============================================================
  Jalankan sekali saat mulai kerja:
    python watcher.py

  Script ini memantau SATU file Excel dengan banyak sheet
  (satu sheet per tim) yang tersync dari SharePoint via OneDrive.
  Setiap sheet berubah (disimpan operator) → hitung agregat
  → simpan ke data/dashboard_data.json → dashboard auto-update.

  DATA YANG DISIMPAN: Hanya agregat (%, total, tren).
  Data individu (Nama, NIP, dll) TIDAK pernah keluar dari Excel.
=============================================================
"""

import json
import os
import sys
import time
import logging
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl belum terinstall. Jalankan: pip install openpyxl")
    sys.exit(1)

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    print("[ERROR] watchdog belum terinstall. Jalankan: pip install watchdog")
    sys.exit(1)

# =============================================================
#  KONFIGURASI — Sesuaikan bagian ini dengan setup Anda
# =============================================================

# Path lengkap ke FILE EXCEL (bukan folder) yang tersync via OneDrive.
# Satu file, banyak sheet — satu sheet per tim.
# Contoh Windows: r"C:\Users\NamaAnda\OneDrive - NamaOrganisasi\Data Sertifikasi\basis data.xlsx"
# Contoh Mac:     "/Users/NamaAnda/Library/CloudStorage/OneDrive-NamaOrg/Data Sertifikasi/basis data.xlsx"
EXCEL_FILE = r"C:\Users\ThinkPad\OneDrive - Kemenkeu\2026\basis data.xlsx"

# Path output JSON (dibaca oleh dashboard — harus di folder yang sama dengan index.html)
OUTPUT_JSON = os.path.join(os.path.dirname(__file__), "data", "dashboard_data.json")

# Mapping nama tim (tampil di dashboard) → nama sheet di file "basis data.xlsx"
# Key   = label yang muncul di dashboard
# Value = nama tab sheet persis di Excel (case-sensitive)
SHEET_MAP = {
    "JFKN":     "JFKN",
    "SAK":      "SAK",
    "AC":       "AC",
    "Beasiswa": "Beasiswa",
    "USKP":     "USKP",
    "PBJ":      "PBJ",
    "PengTes":  "PengTes",
}

# Interval debounce (detik) — hindari proses berulang saat Excel masih dalam proses simpan
DEBOUNCE_SECONDS = 3

# =============================================================
#  KONFIGURASI KOLOM EXCEL PER TIM
# =============================================================

# Nama kolom wajib di Excel (case-insensitive)
COL_STATUS    = "status"       # Nilai: "Lulus" atau "Tidak Lulus"
COL_BATCH     = "batch"
COL_UNIT      = "unit kerja"
COL_PERIODE   = "periode"      # Untuk USKP, PBJ, Tes Lainnya
COL_ANGKATAN  = "angkatan"     # Untuk Beasiswa
COL_PROGRAM   = "program studi"

# Status yang dianggap lulus
STATUS_LULUS = ["lulus", "l", "passed", "p", "ya", "y"]

# =============================================================
#  LOGGING
# =============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(os.path.dirname(__file__), "watcher.log"), encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)

# =============================================================
#  FUNGSI BACA & PROSES EXCEL
# =============================================================

def get_col_index(headers: list, target: str) -> int:
    """Cari index kolom berdasarkan nama (case-insensitive). Return -1 jika tidak ditemukan."""
    target = target.lower().strip()
    for i, h in enumerate(headers):
        if h and str(h).lower().strip() == target:
            return i
    return -1


def process_sheet(ws, team_name: str) -> dict:
    """
    Baca satu sheet Excel, hitung agregat.
    TIDAK menyimpan data individu — hanya statistik.
    """
    try:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            log.warning(f"[{team_name}] Sheet kosong.")
            return _empty_team_data(team_name, error="Sheet kosong")

        # Baris pertama = header
        headers   = [str(c).lower().strip() if c else "" for c in rows[0]]
        data_rows = rows[1:]

        idx_status = get_col_index(headers, COL_STATUS)
        idx_batch  = get_col_index(headers, COL_BATCH)
        idx_unit   = get_col_index(headers, COL_UNIT)

        if idx_status == -1:
            log.error(f"[{team_name}] Kolom 'Status' tidak ditemukan. Cek format sheet.")
            return _empty_team_data(team_name, error="Kolom 'Status' tidak ditemukan")

        total      = 0
        lulus      = 0
        batch_stat = {}
        unit_stat  = {}

        for row in data_rows:
            if not any(row):  # skip baris kosong
                continue

            total += 1
            status_val = str(row[idx_status]).lower().strip() if row[idx_status] else ""
            is_lulus   = status_val in STATUS_LULUS
            if is_lulus:
                lulus += 1

            if idx_batch != -1 and row[idx_batch]:
                b = str(row[idx_batch]).strip()
                if b not in batch_stat:
                    batch_stat[b] = {"total": 0, "lulus": 0}
                batch_stat[b]["total"] += 1
                if is_lulus:
                    batch_stat[b]["lulus"] += 1

            if idx_unit != -1 and row[idx_unit]:
                u = str(row[idx_unit]).strip()
                if u not in unit_stat:
                    unit_stat[u] = {"total": 0, "lulus": 0}
                unit_stat[u]["total"] += 1
                if is_lulus:
                    unit_stat[u]["lulus"] += 1

        pct_lulus  = round((lulus / total * 100), 1) if total > 0 else 0
        batch_list = [
            {"batch": k, "total": v["total"], "lulus": v["lulus"],
             "pct": round(v["lulus"] / v["total"] * 100, 1) if v["total"] > 0 else 0}
            for k, v in batch_stat.items()
        ]
        trend_data = batch_list[-12:] if len(batch_list) > 12 else batch_list
        top_units  = sorted(
            [{"unit": k, "total": v["total"], "lulus": v["lulus"],
              "pct": round(v["lulus"] / v["total"] * 100, 1) if v["total"] > 0 else 0}
             for k, v in unit_stat.items()],
            key=lambda x: x["total"], reverse=True
        )[:10]

        log.info(f"[{team_name}] Diproses: {total} peserta, {lulus} lulus ({pct_lulus}%)")

        return {
            "team": team_name, "total": total, "lulus": lulus,
            "tidak_lulus": total - lulus, "pct_lulus": pct_lulus,
            "trend": trend_data, "top_units": top_units, "error": None,
        }

    except Exception as e:
        log.error(f"[{team_name}] Gagal proses sheet: {e}")
        return _empty_team_data(team_name, error=str(e))


def _empty_team_data(team_name: str, error: str = None) -> dict:
    return {
        "team": team_name, "total": 0, "lulus": 0,
        "tidak_lulus": 0, "pct_lulus": 0,
        "trend": [], "top_units": [], "error": error,
    }


def process_all_sheets() -> None:
    """
    Buka satu file Excel, iterasi setiap sheet sesuai SHEET_MAP,
    hitung agregat per tim, simpan ke JSON.
    File Excel hanya dibuka sekali — efisien dan aman.
    """
    if not os.path.exists(EXCEL_FILE):
        log.error(f"File Excel tidak ditemukan: {EXCEL_FILE}")
        log.error("Pastikan OneDrive sudah sync dan path EXCEL_FILE sudah benar.")
        return

    try:
        log.info(f"Membuka: {os.path.basename(EXCEL_FILE)}")
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

        # Tampilkan sheet yang tersedia untuk membantu debugging
        available_sheets = wb.sheetnames
        log.info(f"Sheet tersedia: {available_sheets}")

        teams_data    = []
        total_peserta = 0
        total_lulus   = 0

        for team_name, sheet_name in SHEET_MAP.items():
            if sheet_name in wb.sheetnames:
                ws     = wb[sheet_name]
                result = process_sheet(ws, team_name)
            else:
                log.warning(f"[{team_name}] Sheet '{sheet_name}' tidak ditemukan di Excel.")
                result = _empty_team_data(team_name, error=f"Sheet '{sheet_name}' tidak ditemukan")

            teams_data.append(result)
            total_peserta += result["total"]
            total_lulus   += result["lulus"]

        wb.close()

    except Exception as e:
        log.error(f"Gagal membuka file Excel: {e}")
        return

    avg_lulus = round(total_lulus / total_peserta * 100, 1) if total_peserta > 0 else 0

    output = {
        "generated_at":  datetime.now().strftime("%d %B %Y, %H:%M:%S"),
        "generated_ts":  datetime.now().isoformat(),
        "total_peserta": total_peserta,
        "total_lulus":   total_lulus,
        "avg_pct_lulus": avg_lulus,
        "total_tim":     len(SHEET_MAP),
        "teams":         teams_data,
    }

    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    log.info(f"✓ JSON diperbarui → {OUTPUT_JSON}")
    log.info(f"  Total: {total_peserta} peserta | Avg lulus: {avg_lulus}%")


# =============================================================
#  FILE WATCHER
# =============================================================

class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self):
        self._last_processed = 0  # debounce tracker

    def on_modified(self, event):
        self._handle(event)

    def on_created(self, event):
        self._handle(event)

    def _handle(self, event):
        if event.is_directory:
            return

        # Hanya proses jika file yang berubah adalah EXCEL_FILE kita
        # Normalisasi path untuk perbandingan lintas OS
        changed = os.path.normcase(os.path.abspath(event.src_path))
        target  = os.path.normcase(os.path.abspath(EXCEL_FILE))
        if changed != target:
            return

        now  = time.time()
        last = self._last_processed
        if now - last < DEBOUNCE_SECONDS:
            return
        self._last_processed = now

        log.info(f"Perubahan terdeteksi di {os.path.basename(EXCEL_FILE)} → memproses ulang...")
        time.sleep(1.5)  # tunggu Excel selesai write (OneDrive kadang butuh sedikit waktu)
        process_all_sheets()


# =============================================================
#  MAIN
# =============================================================

def main():
    log.info("=" * 60)
    log.info("  Dashboard Monitoring Terpadu — Watcher Aktif")
    log.info("=" * 60)
    log.info(f"  File Excel : {EXCEL_FILE}")
    log.info(f"  Output JSON: {OUTPUT_JSON}")
    log.info(f"  Sheet map  : {list(SHEET_MAP.values())}")
    log.info("  Tekan Ctrl+C untuk berhenti.")
    log.info("=" * 60)

    if not os.path.exists(EXCEL_FILE):
        log.error(f"File Excel tidak ditemukan: {EXCEL_FILE}")
        log.error("Pastikan OneDrive sudah sync dan path EXCEL_FILE sudah benar.")
        sys.exit(1)

    # Proses semua sheet saat pertama kali dijalankan
    log.info("Memproses data awal dari semua sheet...")
    process_all_sheets()

    # Watch folder induk dari file Excel (watchdog hanya bisa watch folder, bukan file langsung)
    watch_dir = os.path.dirname(os.path.abspath(EXCEL_FILE))
    handler   = ExcelChangeHandler()
    observer  = Observer()
    observer.schedule(handler, path=watch_dir, recursive=False)
    observer.start()

    log.info(f"Watcher aktif — memantau: {watch_dir}")
    log.info("Dashboard akan update otomatis tiap operator simpan Excel.")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Watcher dihentikan.")
        observer.stop()

    observer.join()


if __name__ == "__main__":
    main()
