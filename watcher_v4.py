"""
=============================================================
  Dashboard Monitoring Terpadu — Excel Watcher Script
  File: basis data.xlsx | Sheet: JFKN, SAK, AC, Beasiswa,
        USKP, PBJ, PengTes
=============================================================
  Jalankan sekali saat mulai kerja:
    python watcher.py

  Memantau satu file Excel multi-sheet via OneDrive sync.
  Setiap file disimpan -> hitung agregat per tim ->
  simpan ke data/dashboard_data.json -> dashboard auto-update.

  DATA YANG DISIMPAN: Hanya agregat (%, total, sebaran).
  Data individu (Nama, NIP, dll) TIDAK pernah keluar.
=============================================================
"""

import json
import os
import sys
import time
import logging
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("[ERROR] openpyxl belum terinstall. Jalankan: pip install openpyxl")
    sys.exit(1)

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    print("[ERROR] watchdog belum terinstall. Jalankan: pip install watchdog")
    sys.exit(1)


# =============================================================
#  KONFIGURASI
# =============================================================

EXCEL_FILE  = r"C:\Users\ThinkPad\OneDrive - Kemenkeu\2026\basis data.xlsx"
OUTPUT_JSON = os.path.join(os.path.dirname(__file__), "data", "dashboard_data.json")
SHEET_MAP   = {
    "JFKN":     "JFKN",
    "SAK":      "SAK",
    "AC":       "AC",
    "Beasiswa": "Beasiswa",
    "USKP":     "USKP",
    "PBJ":      "PBJ",
    "PengTes":  "PengTes",
}
DEBOUNCE_SECONDS = 3
NILAI_LULUS      = {"lulus"}


# =============================================================
#  LOGGING
# =============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(
            os.path.join(os.path.dirname(__file__), "watcher.log"),
            encoding="utf-8"
        ),
    ]
)
log = logging.getLogger(__name__)


# =============================================================
#  HELPER
# =============================================================

def col(headers, *candidates):
    for c in candidates:
        t = c.lower().strip()
        for i, h in enumerate(headers):
            if h and str(h).lower().strip() == t:
                return i
    return -1

def val(row, idx):
    if idx == -1 or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()

def is_lulus(v):
    return v.lower() in NILAI_LULUS

def count_dist(items):
    d = defaultdict(int)
    for x in items:
        if x:
            d[x] += 1
    return dict(sorted(d.items(), key=lambda x: x[1], reverse=True))

def to_rows(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return [], []
    headers = [str(c).lower().strip() if c else "" for c in all_rows[0]]
    data    = [r for r in all_rows[1:] if any(v is not None for v in r)]
    return headers, data

def empty(team, error=None):
    return {
        "team": team, "type": "count", "total": 0,
        "lulus": 0, "tidak_lulus": 0, "pct_lulus": 0,
        "detail": {}, "error": error,
    }


# =============================================================
#  PROCESSOR PER TIM
# =============================================================

def process_JFKN(ws):
    """
    Card   : % Kelulusan — kolom 'Kelulusan' (lulus/tidak lulus)
    Detail : sebaran UE1, UE2, filter Jenjang_Target & JF_Target
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("JFKN", "Sheet kosong")

    i_lulus   = col(headers, "kelulusan")
    i_ue1     = col(headers, "ue1")
    i_ue2     = col(headers, "ue2")
    i_jenjang = col(headers, "jenjang_target")
    i_jf      = col(headers, "jf_target")

    if i_lulus == -1:
        return empty("JFKN", "Kolom 'Kelulusan' tidak ditemukan")

    total = lulus = 0
    ue1_d = defaultdict(lambda: {"total": 0, "lulus": 0})
    ue2_d = defaultdict(lambda: {"total": 0, "lulus": 0})
    jenjang_vals, jf_vals = [], []

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl:
            lulus += 1

        u1 = val(r, i_ue1)
        if u1:
            ue1_d[u1]["total"] += 1
            if vl: ue1_d[u1]["lulus"] += 1

        u2 = val(r, i_ue2)
        if u2:
            ue2_d[u2]["total"] += 1
            if vl: ue2_d[u2]["lulus"] += 1

        jn = val(r, i_jenjang)
        if jn: jenjang_vals.append(jn)
        jf = val(r, i_jf)
        if jf: jf_vals.append(jf)

    pct = round(lulus / total * 100, 1) if total > 0 else 0

    def wpct(d):
        return [{"label": k, "total": v["total"], "lulus": v["lulus"],
                 "pct": round(v["lulus"]/v["total"]*100,1) if v["total"] > 0 else 0}
                for k, v in sorted(d.items(), key=lambda x: x[1]["total"], reverse=True)]

    log.info(f"[JFKN] {total} peserta, {lulus} lulus ({pct}%)")
    return {
        "team": "JFKN", "type": "pct",
        "total": total, "lulus": lulus, "tidak_lulus": total - lulus,
        "pct_lulus": pct, "error": None,
        "detail": {
            "ue1":     wpct(ue1_d),
            "ue2":     wpct(ue2_d),
            "jenjang": count_dist(jenjang_vals),
            "jf":      count_dist(jf_vals),
        }
    }


def process_SAK(ws):
    """
    Card   : % Kelulusan — kolom 'kelulusan'
    Detail : tren per certification_period_name + sebaran exam_location_name
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("SAK", "Sheet kosong")

    i_lulus  = col(headers, "kelulusan")
    i_period = col(headers, "certification_period_name")
    i_lokasi = col(headers, "exam_location_name")

    if i_lulus == -1:
        return empty("SAK", "Kolom 'kelulusan' tidak ditemukan")

    total = lulus = 0
    period_stat = defaultdict(lambda: {"total": 0, "lulus": 0})
    lokasi_vals = []

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl: lulus += 1

        p = val(r, i_period)
        if p:
            period_stat[p]["total"] += 1
            if vl: period_stat[p]["lulus"] += 1

        lk = val(r, i_lokasi)
        if lk: lokasi_vals.append(lk)

    pct   = round(lulus / total * 100, 1) if total > 0 else 0
    trend = [{"batch": k, "total": v["total"], "lulus": v["lulus"],
               "pct": round(v["lulus"]/v["total"]*100,1) if v["total"] > 0 else 0}
             for k, v in period_stat.items()][-12:]

    log.info(f"[SAK] {total} peserta, {lulus} lulus ({pct}%)")
    return {
        "team": "SAK", "type": "pct",
        "total": total, "lulus": lulus, "tidak_lulus": total - lulus,
        "pct_lulus": pct, "error": None,
        "detail": {"trend": trend, "lokasi": count_dist(lokasi_vals)}
    }


def process_AC(ws):
    """
    Card   : Jumlah peserta (semua baris = hadir)
    Detail : sebaran Hasil Penilaian Kompetensi + sebaran UE1 + tren per Batch
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("AC", "Sheet kosong")

    i_hasil = col(headers, "hasil penilaian kompetensi")
    i_ue1   = col(headers, "ue1")
    i_batch = col(headers, "batch")

    hasil_vals, ue1_vals = [], []
    batch_stat = defaultdict(lambda: {"total": 0, "optimal": 0, "cukup": 0, "kurang": 0})

    for r in rows:
        h = val(r, i_hasil)
        if h: hasil_vals.append(h)
        u = val(r, i_ue1)
        if u: ue1_vals.append(u)
        b = val(r, i_batch)
        if b:
            batch_stat[b]["total"] += 1
            hl = h.lower()
            if "optimal" in hl and "cukup" not in hl and "kurang" not in hl:
                batch_stat[b]["optimal"] += 1
            elif "cukup" in hl:
                batch_stat[b]["cukup"] += 1
            elif "kurang" in hl:
                batch_stat[b]["kurang"] += 1

    total = len(rows)
    log.info(f"[AC] {total} peserta | {count_dist(hasil_vals)}")
    return {
        "team": "AC", "type": "count",
        "total": total, "lulus": 0, "tidak_lulus": 0,
        "pct_lulus": 0, "error": None,
        "detail": {
            "hasil":      count_dist(hasil_vals),
            "ue1":        count_dist(ue1_vals),
            "batch_stat": [{"batch": k, **v} for k, v in batch_stat.items()],
        }
    }


def process_Beasiswa(ws):
    """
    Card   : Jumlah penerima beasiswa (semua baris)
    Detail : sebaran prodi, UE1, Universitas, Negara, Jenjang, Beasiswa
             filter dropdown: UE1, Jenjang, Negara
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("Beasiswa", "Sheet kosong")

    i_status = col(headers, "status")
    i_prodi  = col(headers, "prodi")
    i_ue1    = col(headers, "ue1")
    i_univ   = col(headers, "universitas")
    i_negara = col(headers, "negara")
    i_jenj   = col(headers, "jenjang")
    i_bea    = col(headers, "beasiswa")

    status_v, prodi_v = [], []
    ue1_v, univ_v     = [], []
    negara_v, jenj_v  = [], []
    bea_v             = []
    rows_detail       = []

    for r in rows:
        st = val(r, i_status); status_v.append(st) if st else None
        pr = val(r, i_prodi);  prodi_v.append(pr)  if pr else None
        u1 = val(r, i_ue1);    ue1_v.append(u1)    if u1 else None
        un = val(r, i_univ);   univ_v.append(un)   if un else None
        ng = val(r, i_negara); negara_v.append(ng) if ng else None
        jn = val(r, i_jenj);   jenj_v.append(jn)   if jn else None
        bv = val(r, i_bea);    bea_v.append(bv)    if bv else None
        rows_detail.append({
            "ue1": u1, "jenjang": jn, "negara": ng,
            "prodi": pr, "universitas": un, "beasiswa": bv, "status": st,
        })

    total = len(rows)
    log.info(f"[Beasiswa] {total} penerima")
    return {
        "team": "Beasiswa", "type": "count",
        "total": total, "lulus": 0, "tidak_lulus": 0,
        "pct_lulus": 0, "error": None,
        "detail": {
            "status":      count_dist(status_v),
            "prodi":       count_dist(prodi_v),
            "ue1":         count_dist(ue1_v),
            "universitas": count_dist(univ_v),
            "negara":      count_dist(negara_v),
            "jenjang":     count_dist(jenj_v),
            "beasiswa":    count_dist(bea_v),
            "filter_options": {
                "ue1":     sorted(set(x for x in ue1_v    if x)),
                "jenjang": sorted(set(x for x in jenj_v   if x)),
                "negara":  sorted(set(x for x in negara_v if x)),
            },
            "rows": rows_detail,
        }
    }


def process_USKP(ws):
    """
    Card   : % Kelulusan — kolom 'kelulusan'
    Detail : tren per batch + sebaran lokasi
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("USKP", "Sheet kosong")

    i_lulus  = col(headers, "kelulusan")
    i_batch  = col(headers, "batch")
    i_lokasi = col(headers, "lokasi")

    if i_lulus == -1:
        return empty("USKP", "Kolom 'kelulusan' tidak ditemukan")

    total = lulus = 0
    batch_stat  = defaultdict(lambda: {"total": 0, "lulus": 0})
    lokasi_vals = []

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl: lulus += 1

        b = val(r, i_batch)
        if b:
            batch_stat[b]["total"] += 1
            if vl: batch_stat[b]["lulus"] += 1

        lk = val(r, i_lokasi)
        if lk: lokasi_vals.append(lk)

    pct   = round(lulus / total * 100, 1) if total > 0 else 0
    trend = [{"batch": k, "total": v["total"], "lulus": v["lulus"],
               "pct": round(v["lulus"]/v["total"]*100,1) if v["total"] > 0 else 0}
             for k, v in batch_stat.items()][-12:]

    log.info(f"[USKP] {total} peserta, {lulus} lulus ({pct}%)")
    return {
        "team": "USKP", "type": "pct",
        "total": total, "lulus": lulus, "tidak_lulus": total - lulus,
        "pct_lulus": pct, "error": None,
        "detail": {"trend": trend, "lokasi": count_dist(lokasi_vals)}
    }


def process_PBJ(ws):
    """
    Card   : % Kelulusan — kolom 'kelulusan'
    Detail : tren per batch + sebaran jenis ujian
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("PBJ", "Sheet kosong")

    i_lulus = col(headers, "kelulusan")
    i_batch = col(headers, "batch")
    i_jenis = col(headers, "jenis")

    if i_lulus == -1:
        return empty("PBJ", "Kolom 'kelulusan' tidak ditemukan")

    total = lulus = 0
    batch_stat = defaultdict(lambda: {"total": 0, "lulus": 0})
    jenis_vals = []

    for r in rows:
        total += 1
        vl = is_lulus(val(r, i_lulus))
        if vl: lulus += 1

        b = val(r, i_batch)
        if b:
            batch_stat[b]["total"] += 1
            if vl: batch_stat[b]["lulus"] += 1

        j = val(r, i_jenis)
        if j: jenis_vals.append(j)

    pct   = round(lulus / total * 100, 1) if total > 0 else 0
    trend = [{"batch": k, "total": v["total"], "lulus": v["lulus"],
               "pct": round(v["lulus"]/v["total"]*100,1) if v["total"] > 0 else 0}
             for k, v in batch_stat.items()][-12:]

    log.info(f"[PBJ] {total} peserta, {lulus} lulus ({pct}%)")
    return {
        "team": "PBJ", "type": "pct",
        "total": total, "lulus": lulus, "tidak_lulus": total - lulus,
        "pct_lulus": pct, "error": None,
        "detail": {"trend": trend, "jenis": count_dist(jenis_vals)}
    }


def process_PengTes(ws):
    """
    Card   : Jumlah peserta (semua baris = hadir)
    Detail : sebaran jenis ujian (TPA/TBI/lainnya) + sebaran UE1
             (kelulusan tidak ditampilkan)
    """
    headers, rows = to_rows(ws)
    if not rows:
        return empty("PengTes", "Sheet kosong")

    i_ujian = col(headers, "ujian")
    i_ue1   = col(headers, "ue1")
    i_batch = col(headers, "batch")

    ujian_vals, ue1_vals = [], []
    batch_stat = defaultdict(int)

    for r in rows:
        u = val(r, i_ujian)
        if u: ujian_vals.append(u)
        u1 = val(r, i_ue1)
        if u1: ue1_vals.append(u1)
        b = val(r, i_batch)
        if b: batch_stat[b] += 1

    total = len(rows)
    log.info(f"[PengTes] {total} peserta | ujian: {count_dist(ujian_vals)}")
    return {
        "team": "PengTes", "type": "count",
        "total": total, "lulus": 0, "tidak_lulus": 0,
        "pct_lulus": 0, "error": None,
        "detail": {
            "ujian":      count_dist(ujian_vals),
            "ue1":        count_dist(ue1_vals),
            "batch_stat": [{"batch": k, "total": v} for k, v in batch_stat.items()],
        }
    }


PROCESSORS = {
    "JFKN":     process_JFKN,
    "SAK":      process_SAK,
    "AC":       process_AC,
    "Beasiswa": process_Beasiswa,
    "USKP":     process_USKP,
    "PBJ":      process_PBJ,
    "PengTes":  process_PengTes,
}


# =============================================================
#  PROSES SEMUA SHEET
# =============================================================

def process_all_sheets():
    if not os.path.exists(EXCEL_FILE):
        log.error(f"File tidak ditemukan: {EXCEL_FILE}")
        return

    try:
        log.info(f"Membuka: {os.path.basename(EXCEL_FILE)}")
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        log.info(f"Sheet tersedia: {wb.sheetnames}")

        teams_data = []
        for team_name, sheet_name in SHEET_MAP.items():
            if sheet_name not in wb.sheetnames:
                log.warning(f"[{team_name}] Sheet '{sheet_name}' tidak ditemukan.")
                teams_data.append(empty(team_name, f"Sheet '{sheet_name}' tidak ditemukan"))
                continue
            processor = PROCESSORS.get(team_name)
            if processor:
                teams_data.append(processor(wb[sheet_name]))
            else:
                teams_data.append(empty(team_name, "Processor tidak ditemukan"))

        wb.close()

    except Exception as e:
        log.error(f"Gagal membuka file Excel: {e}")
        return

    pct_teams     = [t for t in teams_data if t["type"] == "pct"]
    total_peserta = sum(t["total"] for t in pct_teams)
    total_lulus   = sum(t["lulus"]  for t in pct_teams)
    avg_lulus     = round(total_lulus / total_peserta * 100, 1) if total_peserta > 0 else 0

    output = {
        "generated_at":  datetime.now().strftime("%d %B %Y, %H:%M:%S"),
        "generated_ts":  datetime.now().isoformat(),
        "total_peserta": total_peserta,
        "total_lulus":   total_lulus,
        "avg_pct_lulus": avg_lulus,
        "total_tim":     len(SHEET_MAP),
        "teams":         teams_data,
    }

    os.makedirs(os.path.dirname(OUTPUT_JSON), exist_ok=True)
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    log.info(f"JSON diperbarui -> {OUTPUT_JSON}")
    log.info(f"  Ringkasan pct-tim: {total_peserta} peserta | avg lulus {avg_lulus}%")


# =============================================================
#  FILE WATCHER
# =============================================================

class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self):
        self._last = 0

    def on_modified(self, event):
        self._handle(event)

    def on_created(self, event):
        self._handle(event)

    def _handle(self, event):
        if event.is_directory:
            return
        if os.path.normcase(os.path.abspath(event.src_path)) != \
           os.path.normcase(os.path.abspath(EXCEL_FILE)):
            return
        now = time.time()
        if now - self._last < DEBOUNCE_SECONDS:
            return
        self._last = now
        log.info("Perubahan terdeteksi -> memproses ulang...")
        time.sleep(1.5)
        process_all_sheets()


# =============================================================
#  MAIN
# =============================================================

def main():
    log.info("=" * 60)
    log.info("  Dashboard Monitoring Terpadu — Watcher Aktif")
    log.info("=" * 60)
    log.info(f"  File  : {EXCEL_FILE}")
    log.info(f"  Output: {OUTPUT_JSON}")
    log.info(f"  Sheet : {list(SHEET_MAP.values())}")
    log.info("  Ctrl+C untuk berhenti.")
    log.info("=" * 60)

    if not os.path.exists(EXCEL_FILE):
        log.error(f"File tidak ditemukan: {EXCEL_FILE}")
        sys.exit(1)

    log.info("Memproses data awal...")
    process_all_sheets()

    watch_dir = os.path.dirname(os.path.abspath(EXCEL_FILE))
    handler   = ExcelChangeHandler()
    observer  = Observer()
    observer.schedule(handler, path=watch_dir, recursive=False)
    observer.start()
    log.info(f"Watcher aktif — memantau: {watch_dir}")

    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        log.info("Watcher dihentikan.")
        observer.stop()
    observer.join()


if __name__ == "__main__":
    main()
