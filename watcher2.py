"""
=============================================================
  Dashboard Monitoring Terpadu — Excel Watcher Script v2
  Arsitektur 2-layer:
    - summary.json              (ringan, untuk halaman utama)
    - detail_{tim}_{thn}.json   (lazy-load, per tim per tahun)
=============================================================
"""

import json, os, sys, time, logging, re
from datetime import datetime
from collections import defaultdict, OrderedDict

try:
    import openpyxl
except ImportError:
    print("[ERROR] pip install openpyxl"); sys.exit(1)

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    print("[ERROR] pip install watchdog"); sys.exit(1)


# =============================================================
#  KONFIGURASI
# =============================================================

EXCEL_FILE  = r"C:\Users\agungadhi\OneDrive - Kemenkeu\2026\basis data.xlsx"
OUTPUT_DIR  = os.path.join(os.path.dirname(__file__), "data")
SHEET_MAP   = {
    "JFKN":      "JFKN",
    "SAK":       "SAK",
    "AC":        "AC",
    "Beasiswa":  "Beasiswa",
    "USKP":      "USKP",
    "PBJ":       "PBJ",
    "Tes Lain":  "Tes Lain",
}

# Sub-exam sheets yang ditampilkan di dalam card JFKN
JFKN_SUB_SHEETS = {
    "PKAPK APBN": {"sheet": "PKAPK APBN", "col_hasil": "hasil ukom",        "label_lulus": "Lulus",              "label_tidak": "Tidak Lulus"},
    "Pelelang":   {"sheet": "Pelelang",   "col_hasil": "hasil ukom teknis",  "label_lulus": "Lulus Ukom Teknis",  "label_tidak": "Tidak Lulus Ukom Teknis"},
    "AKPD":       {"sheet": "AKPD",       "col_hasil": "hasil ukom teknis",  "label_lulus": "Lulus Ukom Teknis",  "label_tidak": "Tidak Lulus Ukom Teknis"},
    "AA":         {"sheet": "AA",         "col_hasil": "hasil",              "label_lulus": "Lulus",              "label_tidak": "Tidak Lulus"},
}
DEBOUNCE_SECONDS = 3
NILAI_LULUS      = {"lulus", "direkomendasikan"}

MONTH_EN = {
    "january":1,"february":2,"march":3,"april":4,"may":5,"june":6,
    "july":7,"august":8,"september":9,"october":10,"november":11,"december":12,
    "jan":1,"feb":2,"mar":3,"apr":4,"jun":6,"jul":7,"aug":8,
    "sep":9,"oct":10,"nov":11,"dec":12,
}


# =============================================================
#  LOGGING
# =============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(os.path.join(os.path.dirname(__file__), "watcher.log"), encoding="utf-8"),
    ]
)
log = logging.getLogger(__name__)


# =============================================================
#  HELPERS
# =============================================================

def col(headers, *candidates):
    for c in candidates:
        t = c.lower().strip()
        for i, h in enumerate(headers):
            if h and str(h).lower().strip() == t: return i
    return -1

def val(row, idx):
    if idx == -1 or idx >= len(row) or row[idx] is None: return ""
    return str(row[idx]).strip()

def is_lulus(v): return v.lower() in NILAI_LULUS

def count_dist(items):
    d = defaultdict(int)
    for x in items:
        if x: d[x] += 1
    return dict(sorted(d.items(), key=lambda x: x[1], reverse=True))

def to_rows(ws):
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows: return [], []
    headers = [str(c).lower().strip() if c else "" for c in all_rows[0]]
    data = [r for r in all_rows[1:] if any(v is not None for v in r)]
    return headers, data

def extract_month_year(value):
    if value is None: return None, None
    if hasattr(value, 'month'): return value.month, value.year
    if isinstance(value, (int, float)) and 30000 < value < 60000:
        from datetime import date, timedelta
        serial = int(value)
        if serial > 59: serial -= 1
        d = date(1899, 12, 31) + timedelta(days=serial)
        return d.month, d.year
    s = str(value).strip()
    if not s or s.lower() == 'none': return None, None
    sl = s.lower()
    if sl in MONTH_EN: return MONTH_EN[sl], None
    for name, num in MONTH_EN.items():
        if len(name) >= 3 and name in sl:
            m = re.search(r'\b(20\d{2})\b', s)
            yr = int(m.group(1)) if m else None
            return num, yr
    for sep in ['/', '-', '.']:
        parts = s.split(sep)
        if len(parts) == 3:
            try:
                nums = [int(p) for p in parts]
                if nums[0] > 31: return nums[1], nums[0]
                if nums[2] > 31: return nums[1], nums[2]
            except: pass
    return None, None

def add_month_year(months_dict, value, lulus=False):
    m, y = extract_month_year(value)
    if m is None: return
    if y is None: y = datetime.now().year
    if y not in months_dict: months_dict[y] = {}
    if m not in months_dict[y]: months_dict[y][m] = {"total": 0, "lulus": 0}
    months_dict[y][m]["total"] += 1
    if lulus: months_dict[y][m]["lulus"] += 1

def safe_slug(name):
    return name.lower().replace(' ', '_')

def _empty(team, error=None):
    return {
        "team": team, "type": "count", "total": 0,
        "lulus": 0, "tidak_lulus": 0, "pct_lulus": 0,
        "total_angkatan": 0, "total_hadir": 0,
        "months_by_year": {}, "error": error,
    }


# =============================================================
#  BACA SHEET OUTPUT (TARGETS)
# =============================================================

def read_targets(wb):
    targets = {}
    sheet_name = next((s for s in wb.sheetnames if s.lower() == 'output'), None)
    if not sheet_name:
        log.warning("Sheet 'output' tidak ditemukan.")
        return targets
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return targets
    headers = [str(c).strip() if c else "" for c in rows[0]]
    col_ujian = next((i for i, h in enumerate(headers) if h.lower() == 'ujian'), -1)
    if col_ujian == -1: return targets
    year_cols = {h: i for i, h in enumerate(headers) if h.isdigit() and len(h) == 4}
    for row in rows[1:]:
        if not any(v is not None for v in row): continue
        team = str(row[col_ujian]).strip() if row[col_ujian] else ""
        if not team: continue
        targets[team] = {}
        for yr, idx in year_cols.items():
            if idx < len(row) and row[idx] is not None:
                try: targets[team][int(yr)] = int(row[idx])
                except: pass
    log.info(f"Targets: {targets}")
    return targets


# =============================================================
#  PROCESSORS — return (summary_dict, {year: [rows]})
# =============================================================

def process_JFKN(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("JFKN", "Sheet kosong"), {}
    i_lulus = col(headers, "kelulusan")
    i_ue1 = col(headers, "ue1"); i_ue2 = col(headers, "ue2")
    i_jenjang = col(headers, "jenjang_target"); i_jf = col(headers, "jf_target")
    i_tgl = col(headers, "tglukom", "tgl_ukom"); i_hadir = col(headers, "kehadiran")
    i_status = col(headers, "status")
    if i_lulus == -1: return _empty("JFKN", "Kolom 'Kelulusan' tidak ditemukan"), {}

    total = lulus_cnt = n_hadir = n_belum = 0
    months_by_year = {}; rows_by_year = defaultdict(list)

    for r in rows:
        raw_kel = val(r, i_lulus)
        if raw_kel and '/' in raw_kel: continue
        total += 1
        is_hadir = False
        if i_hadir != -1:
            h = val(r, i_hadir).lower().strip()
            if h == "hadir": n_hadir += 1; is_hadir = True
            else: n_belum += 1
        else: is_hadir = True; n_hadir += 1
        vl = is_hadir and is_lulus(raw_kel) if raw_kel else False
        if vl: lulus_cnt += 1
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        m, y = extract_month_year(raw_tgl)
        if y is None: y = datetime.now().year
        if is_hadir: add_month_year(months_by_year, raw_tgl, lulus=vl)
        rows_by_year[y].append({
            "hadir": is_hadir, "lulus": vl, "status": val(r, i_status),
            "ue1": val(r, i_ue1), "ue2": val(r, i_ue2),
            "jenjang": val(r, i_jenjang), "jf": val(r, i_jf),
            "month": m, "year": y,
        })

    pct = round(lulus_cnt/n_hadir*100, 1) if n_hadir > 0 else 0
    hby = {yr: sum(1 for r in rl if r["hadir"]) for yr, rl in rows_by_year.items()}
    fby = {}
    for yr, rl in rows_by_year.items():
        fby[str(yr)] = {
            "terdaftar": len(rl), "mengikuti": sum(1 for r in rl if r["hadir"]),
            "direkomendasikan": sum(1 for r in rl if r["lulus"]),
            "belum_diproses": sum(1 for r in rl if not r["hadir"]),
        }
    log.info(f"[JFKN] {total} terdaftar, {n_hadir} mengikuti, {lulus_cnt} lulus ({pct}%)")
    return {
        "team": "JFKN", "type": "pct", "total": total, "total_hadir": n_hadir,
        "lulus": lulus_cnt, "tidak_lulus": n_hadir - lulus_cnt, "dalam_proses": n_belum,
        "pct_lulus": pct, "total_angkatan": 0,
        "targets": targets.get("JFKN", {}), "months_by_year": months_by_year,
        "hadir_by_year": hby, "funnel_by_year": fby, "error": None,
    }, dict(rows_by_year)


def process_SAK(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("SAK", "Sheet kosong"), {}
    i_lulus = col(headers, "kelulusan")
    i_period = col(headers, "certification_period_name")
    i_mon = col(headers, "certification_period_month")
    i_yr_col = col(headers, "certification_period_year")
    i_tgl = col(headers, "tgl_ujian", "tgl_ukom", "tanggal_ujian")
    i_lokasi = col(headers, "exam_location_name")
    i_vdok = col(headers, "verifikasi dokumen"); i_vpay = col(headers, "verifikasi pembayaran")
    i_hadir = col(headers, "kehadiran")
    i_tingkat = col(headers, "tingkat", "level")
    if i_lulus == -1: return _empty("SAK", "Kolom 'kelulusan' tidak ditemukan"), {}

    total = lulus_cnt = n_vdok = n_vpay = n_hadir = 0
    months_by_year = {}; rows_by_year = defaultdict(list)

    for r in rows:
        total += 1; is_hadir = False
        if i_hadir != -1:
            h = val(r, i_hadir).lower()
            if "hadir" in h and "tidak" not in h: n_hadir += 1; is_hadir = True
        vl = is_hadir and is_lulus(val(r, i_lulus))
        if vl: lulus_cnt += 1
        is_vdok = i_vdok != -1 and bool(val(r, i_vdok))
        is_vpay = i_vpay != -1 and bool(val(r, i_vpay))
        if is_vdok: n_vdok += 1
        if is_vpay: n_vpay += 1
        # Tahun: utamakan certification_period_year, fallback ke tgl_ujian
        raw_yr  = r[i_yr_col] if i_yr_col != -1 and i_yr_col < len(r) else None
        raw_mon = r[i_mon]    if i_mon    != -1 and i_mon    < len(r) else None
        raw_tgl = r[i_tgl]    if i_tgl    != -1 and i_tgl    < len(r) else None
        y = None
        if raw_yr:
            try: y = int(str(raw_yr).strip())
            except: pass
        if y is None and raw_tgl:
            _, y = extract_month_year(raw_tgl)
        if y is None and raw_mon:
            _, y2 = extract_month_year(raw_mon)
            if y2: y = y2
        if y is None: y = datetime.now().year
        # Bulan: utamakan certification_period_month (integer langsung)
        m = None
        if raw_mon is not None:
            try: m = int(str(raw_mon).strip())
            except: pass
        if m is None:
            m, _ = extract_month_year(raw_tgl)
        # Fallback: parse dari certification_period_name (misal "2026 - Februari")
        if m is None and i_period != -1 and i_period < len(r) and r[i_period]:
            _BULAN = {
                'januari':1,'februari':2,'maret':3,'april':4,'mei':5,'juni':6,
                'juli':7,'agustus':8,'september':9,'oktober':10,'november':11,'desember':12
            }
            _parts = str(r[i_period]).lower().replace('-',' ').split()
            m = next((_BULAN[p] for p in _parts if p in _BULAN), None)
        if is_hadir and m is not None:
            if y not in months_by_year: months_by_year[y] = {}
            if m not in months_by_year[y]: months_by_year[y][m] = {"total": 0, "lulus": 0}
            months_by_year[y][m]["total"] += 1
            if vl: months_by_year[y][m]["lulus"] += 1
        rows_by_year[y].append({
            "hadir": is_hadir, "lulus": vl, "verif_dok": is_vdok, "verif_pay": is_vpay,
            "batch": val(r, i_period), "lokasi": val(r, i_lokasi),
            "tingkat": val(r, i_tingkat), "month": m, "year": y,
        })

    pct = round(lulus_cnt/n_hadir*100, 1) if n_hadir > 0 else 0
    hby = {yr: sum(1 for r in rl if r["hadir"]) for yr, rl in rows_by_year.items()}
    log.info(f"[SAK] {total} pendaftar, {n_hadir} hadir, {lulus_cnt} lulus ({pct}%)")
    for yr, rl in sorted(rows_by_year.items()):
        log.info(f"  [SAK] {yr}: {len(rl)} baris, {sum(1 for r in rl if r['hadir'])} hadir")
    fby = {}
    for yr, rl in rows_by_year.items():
        fby[str(yr)] = {
            "pendaftar": len(rl), "verif_dokumen": sum(1 for r in rl if r["verif_dok"]),
            "verif_pembayaran": sum(1 for r in rl if r["verif_pay"]),
            "hadir": sum(1 for r in rl if r["hadir"]), "lulus": sum(1 for r in rl if r["lulus"]),
        }
    log.info(f"[SAK] {total} pendaftar, {n_hadir} hadir, {lulus_cnt} lulus ({pct}%)")
    return {
        "team": "SAK", "type": "pct", "total": total, "total_hadir": n_hadir,
        "lulus": lulus_cnt, "tidak_lulus": n_hadir - lulus_cnt, "pct_lulus": pct,
        "total_angkatan": len(set(r["batch"] for rl in rows_by_year.values() for r in rl if r["hadir"] and r["batch"])),
        "targets": targets.get("SAK", {}), "months_by_year": months_by_year,
        "hadir_by_year": hby, "funnel_by_year": fby, "error": None,
    }, dict(rows_by_year)


def process_AC(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("AC", "Sheet kosong"), {}
    i_hasil   = col(headers, "hasil penilaian kompetensi")
    i_jabatan = col(headers, "jabatan saat ac")
    i_batch   = col(headers, "batch")
    i_tgl     = col(headers, "tanggal ac")

    months_by_year = {}
    batch_by_year = {}; jabatan_by_year = {}; hasil_by_year = {}

    for r in rows:
        h = val(r, i_hasil); jbt = val(r, i_jabatan); b = val(r, i_batch)
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        _, yr = extract_month_year(raw_tgl)
        if yr is None: yr = datetime.now().year
        hl = h.lower() if h else ""
        if b:
            if yr not in batch_by_year: batch_by_year[yr] = OrderedDict()
            if b not in batch_by_year[yr]: batch_by_year[yr][b] = {"total":0,"optimal":0,"cukup":0,"kurang":0}
            batch_by_year[yr][b]["total"] += 1
            if "optimal" in hl and "cukup" not in hl and "kurang" not in hl: batch_by_year[yr][b]["optimal"] += 1
            elif "cukup" in hl: batch_by_year[yr][b]["cukup"] += 1
            elif "kurang" in hl: batch_by_year[yr][b]["kurang"] += 1
        if jbt:
            if yr not in jabatan_by_year: jabatan_by_year[yr] = {}
            jabatan_by_year[yr][jbt] = jabatan_by_year[yr].get(jbt, 0) + 1
        if h:
            if yr not in hasil_by_year: hasil_by_year[yr] = {}
            hasil_by_year[yr][h] = hasil_by_year[yr].get(h, 0) + 1
        add_month_year(months_by_year, raw_tgl)

    total = len(rows)
    total_batches = set()
    for bmap in batch_by_year.values():
        total_batches.update(bmap.keys())
    hby = {yr: sum(v["total"] for v in bmap.values()) for yr, bmap in batch_by_year.items()}
    log.info(f"[AC] {total} peserta, tahun: {sorted(hby.keys())}")
    return {
        "team": "AC", "type": "count", "total": total, "total_hadir": total,
        "lulus": 0, "tidak_lulus": 0, "pct_lulus": 0,
        "total_angkatan": len(total_batches),
        "targets": targets.get("AC", {}), "months_by_year": months_by_year,
        "hadir_by_year": hby,
        "detail": {
            "hasil_by_year":   {str(yr): d for yr, d in hasil_by_year.items()},
            "batch_by_year":   {str(yr): [{"batch":k,**v} for k,v in bmap.items()] for yr, bmap in batch_by_year.items()},
            "jabatan_by_year": {str(yr): d for yr, d in jabatan_by_year.items()},
        },
        "error": None,
    }, {}

def process_Beasiswa(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("Beasiswa", "Sheet kosong"), {}
    i_status = col(headers, "status"); i_prodi = col(headers, "prodi")
    i_ue1 = col(headers, "ue1"); i_univ = col(headers, "universitas")
    i_negara = col(headers, "negara"); i_jenj = col(headers, "jenjang")
    i_bea = col(headers, "beasiswa")
    i_ptud = col(headers, "ptud")           # PTUD / Non PTUD
    i_lokasi = col(headers, "lokasi")       # Dalam Negeri / Luar Negeri
    i_tgl = col(headers, "tgl_regis", "tgl regis", "tanggal registrasi")

    months_by_year = {}; all_rows = []
    for r in rows:
        st = val(r, i_status); pr = val(r, i_prodi); u1 = val(r, i_ue1)
        un = val(r, i_univ); ng = val(r, i_negara); jn = val(r, i_jenj); bv = val(r, i_bea)
        ptud_val = val(r, i_ptud)       # "PTUD" atau "Non PTUD"
        lokasi_val = val(r, i_lokasi)   # "Dalam Negeri" atau "Luar Negeri"
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        m, y = extract_month_year(raw_tgl)
        add_month_year(months_by_year, raw_tgl)
        all_rows.append({
            "ue1":u1, "jenjang":jn, "negara":ng, "prodi":pr,
            "universitas":un, "beasiswa":bv, "status":st,
            "ptud": ptud_val, "lokasi": lokasi_val,
            "month":m, "year":y,
        })

    aktif = [r for r in all_rows if (r["status"] or "").strip() == "Aktif"]
    selesai = [r for r in all_rows if (r["status"] or "").strip() == "Selesai"]
    combined = aktif + selesai

    def cross(key):
        """Count aktif vs selesai per value of key"""
        d = {}
        for r in combined:
            v = r.get(key, "")
            if not v: continue
            if v not in d: d[v] = {"aktif":0,"selesai":0}
        for r in aktif:
            v = r.get(key, "")
            if v and v in d: d[v]["aktif"] += 1
        for r in selesai:
            v = r.get(key, "")
            if v:
                if v not in d: d[v] = {"aktif":0,"selesai":0}
                d[v]["selesai"] += 1
        return dict(sorted(d.items(), key=lambda x:-(x[1]["aktif"]+x[1]["selesai"])))

    # DN/LN dari kolom "lokasi" (Dalam Negeri / Luar Negeri)
    # Fallback ke kolom "negara" jika kolom lokasi belum terisi
    def is_dn(r):
        lok = (r.get("lokasi") or "").strip().lower()
        if lok:
            return "dalam" in lok  # "Dalam Negeri"
        # fallback: negara Indonesia = DN
        return (r.get("negara") or "").strip() == "Indonesia"

    dn_a = sum(1 for r in aktif if is_dn(r))
    ln_a = len(aktif) - dn_a
    dn_s = sum(1 for r in selesai if is_dn(r))
    ln_s = len(selesai) - dn_s

    # PTUD: filter hanya yang PTUD, lalu cross per universitas
    ptud_rows = [r for r in combined if (r.get("ptud") or "").strip().upper() == "PTUD"]
    ptud_data = {}
    for r in ptud_rows:
        univ = r.get("universitas", "")
        if not univ: continue
        if univ not in ptud_data: ptud_data[univ] = {"aktif":0,"selesai":0}
        st = (r.get("status") or "").strip()
        if st == "Aktif": ptud_data[univ]["aktif"] += 1
        elif st == "Selesai": ptud_data[univ]["selesai"] += 1
    ptud_sorted = dict(sorted(ptud_data.items(), key=lambda x:-(x[1]["aktif"]+x[1]["selesai"])))

    # PTUD vs Non PTUD — agregat aktif/selesai
    ptud_status = {"PTUD": {"aktif":0,"selesai":0}, "Non PTUD": {"aktif":0,"selesai":0}}
    for r in combined:
        pk = (r.get("ptud") or "").strip()
        if pk not in ptud_status: continue
        st = (r.get("status") or "").strip()
        if st == "Aktif": ptud_status[pk]["aktif"] += 1
        elif st == "Selesai": ptud_status[pk]["selesai"] += 1

    log.info(f"[Beasiswa] {len(all_rows)} total, {len(aktif)} aktif, {len(selesai)} selesai, {len(ptud_rows)} PTUD")
    return {
        "team": "Beasiswa", "type": "count",
        "total": len(aktif), "total_hadir": len(aktif), "total_selesai": len(selesai),
        "lulus": 0, "tidak_lulus": 0, "pct_lulus": 0, "total_angkatan": 0,
        "targets": targets.get("Beasiswa", {}), "months_by_year": months_by_year,
        "detail": {
            "dn_ln": {
                "Dalam Negeri": {"aktif": dn_a, "selesai": dn_s},
                "Luar Negeri":  {"aktif": ln_a, "selesai": ln_s},
            },
            "jenjang": cross("jenjang"),
            "provider": cross("beasiswa"),
            "ptud": ptud_sorted,
            "ptud_status": ptud_status,
            "ue1": count_dist([r["ue1"] for r in combined if r["ue1"]]),
            "prodi": count_dist([r["prodi"] for r in combined if r["prodi"]]),
        },
        "error": None,
    }, {}


def process_USKP(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("USKP", "Sheet kosong"), {}
    i_lulus = col(headers, "kelulusan"); i_batch = col(headers, "batch")
    i_lokasi = col(headers, "lokasi"); i_verif = col(headers, "verifikasi")
    i_hadir = col(headers, "kehadiran"); i_tgl = col(headers, "tgl_ujian")
    i_tingkat = col(headers, "tingkat")
    if i_lulus == -1: return _empty("USKP", "Kolom 'kelulusan' tidak ditemukan"), {}

    total = lulus_cnt = n_hadir = n_lolos = 0
    months_by_year = {}; rows_by_year = defaultdict(list)
    # kelulusan_dist: {value_asli: count} untuk semua peserta hadir
    kelulusan_dist_by_year = defaultdict(lambda: defaultdict(int))

    for r in rows:
        total += 1; is_hadir = False
        raw_had = val(r, i_hadir).lower().strip() if i_hadir != -1 else ''
        raw_ver = val(r, i_verif).lower().strip() if i_verif != -1 else ''
        raw_lul_raw = val(r, i_lulus).strip()          # nilai asli, pertahankan casing
        raw_lul     = raw_lul_raw.lower()
        is_lolos = 'lolos' in raw_ver and 'tidak' not in raw_ver
        if is_lolos: n_lolos += 1
        if raw_had:
            if 'hadir' in raw_had and 'tidak' not in raw_had: n_hadir += 1; is_hadir = True
        elif is_lolos: n_hadir += 1; is_hadir = True
        vl = is_hadir and raw_lul == 'lulus'
        if vl: lulus_cnt += 1
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        # tgl_ujian berisi tahun saja (mis. 2025) — baca langsung sebagai int
        y = None
        if raw_tgl is not None:
            try: y = int(str(raw_tgl).strip())
            except: pass
        if y is None: y = datetime.now().year
        if is_hadir:
            add_month_year(months_by_year, raw_tgl, lulus=vl)
            kel_key = raw_lul_raw if raw_lul_raw else '(kosong)'
            kelulusan_dist_by_year[y][kel_key] += 1
        rows_by_year[y].append({
            "hadir": is_hadir, "lolos_verif": is_lolos, "lulus": vl,
            "kelulusan": raw_lul_raw,   # nilai asli — untuk filter & tampilan
            "batch": val(r, i_batch), "lokasi": val(r, i_lokasi),
            "tingkat": val(r, i_tingkat), "month": None, "year": y,
        })

    pct = round(lulus_cnt/n_hadir*100, 1) if n_hadir > 0 else 0
    hby = {yr: sum(1 for r in rl if r["hadir"]) for yr, rl in rows_by_year.items()}
    fby = {}
    for yr, rl in rows_by_year.items():
        fby[str(yr)] = {
            "pendaftar": len(rl), "lolos_verifikasi": sum(1 for r in rl if r["lolos_verif"]),
            "hadir": sum(1 for r in rl if r["hadir"]), "lulus": sum(1 for r in rl if r["lulus"]),
        }
    log.info(f"[USKP] {total} pendaftar, {n_hadir} hadir, {lulus_cnt} lulus ({pct}%)")
    for yr, dist in sorted(kelulusan_dist_by_year.items()):
        log.info(f"  [USKP] {yr} kelulusan: {dict(dist)}")
    return {
        "team": "USKP", "type": "pct", "total": total, "total_hadir": n_hadir,
        "lulus": lulus_cnt, "tidak_lulus": n_hadir - lulus_cnt,
        "kelulusan_dist_by_year": {str(yr): dict(d) for yr, d in kelulusan_dist_by_year.items()},
        "pct_lulus": pct,
        "total_angkatan": len(set(r["batch"] for rl in rows_by_year.values() for r in rl if r["hadir"] and r["batch"])),
        "targets": targets.get("USKP", {}), "months_by_year": months_by_year,
        "hadir_by_year": hby, "funnel_by_year": fby, "error": None,
    }, dict(rows_by_year)


def process_PBJ(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("PBJ", "Sheet kosong"), {}
    i_lulus = col(headers, "kelulusan"); i_batch = col(headers, "batch")
    i_level = col(headers, "level")
    i_ue1   = col(headers, "ue1")
    i_unit  = col(headers, "unit kerja", "unit_kerja")
    i_lokasi= col(headers, "lokasi")
    i_verif = col(headers, "verifikasi"); i_hadir = col(headers, "kehadiran")
    i_tgl   = col(headers, "tgl_ujian")
    if i_lulus == -1: return _empty("PBJ", "Kolom 'kelulusan' tidak ditemukan"), {}

    total = lulus_cnt = n_hadir = n_lolos = n_belum_ada = 0
    months_by_year = {}; rows_by_year = defaultdict(list)

    for r in rows:
        total += 1; is_hadir = False
        if i_hadir != -1:
            h = val(r, i_hadir).lower()
            if "hadir" in h and "tidak" not in h: n_hadir += 1; is_hadir = True
        raw_kel_raw = val(r, i_lulus).strip()          # nilai asli, pertahankan casing
        raw_kel     = raw_kel_raw.lower()
        is_belum_ada = is_hadir and raw_kel == "belum ada"
        vl = is_hadir and is_lulus(raw_kel_raw)
        if vl: lulus_cnt += 1
        if is_belum_ada: n_belum_ada += 1
        is_lolos = False
        if i_verif != -1:
            v = val(r, i_verif).lower()
            if "lolos" in v and "tidak" not in v: n_lolos += 1; is_lolos = True
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        m, y = extract_month_year(raw_tgl)
        if y is None: y = datetime.now().year
        if is_hadir: add_month_year(months_by_year, raw_tgl, lulus=vl)
        rows_by_year[y].append({
            "hadir": is_hadir, "lolos_verif": is_lolos, "lulus": vl,
            "belum_ada": is_belum_ada,
            "kelulusan": raw_kel_raw,   # nilai asli — untuk tampilan & filter
            "batch": val(r, i_batch), "level": val(r, i_level),
            "ue1": val(r, i_ue1), "unit_kerja": val(r, i_unit),
            "lokasi": val(r, i_lokasi), "month": m, "year": y,
        })

    # tidak_lulus = hadir, bukan lulus, dan bukan belum_ada
    tidak_lulus_cnt = n_hadir - lulus_cnt - n_belum_ada
    pct = round(lulus_cnt/n_hadir*100, 1) if n_hadir > 0 else 0
    hby = {yr: sum(1 for r in rl if r["hadir"]) for yr, rl in rows_by_year.items()}
    fby = {}
    for yr, rl in rows_by_year.items():
        fby[str(yr)] = {
            "pendaftar": len(rl), "lolos_verifikasi": sum(1 for r in rl if r["lolos_verif"]),
            "hadir": sum(1 for r in rl if r["hadir"]), "lulus": sum(1 for r in rl if r["lulus"]),
            "belum_ada": sum(1 for r in rl if r["belum_ada"]),
        }
    log.info(f"[PBJ] {total} pendaftar, {n_hadir} hadir, {lulus_cnt} lulus, {n_belum_ada} belum ada ({pct}%)")
    return {
        "team": "PBJ", "type": "pct", "total": total, "total_hadir": n_hadir,
        "lulus": lulus_cnt, "tidak_lulus": tidak_lulus_cnt,
        "belum_ada": n_belum_ada, "pct_lulus": pct,
        "total_angkatan": len(set(r["batch"] for rl in rows_by_year.values() for r in rl if r["hadir"] and r["batch"])),
        "targets": targets.get("PBJ", {}), "months_by_year": months_by_year,
        "hadir_by_year": hby, "funnel_by_year": fby, "error": None,
    }, dict(rows_by_year)


def process_TesTes(ws, targets):
    headers, rows = to_rows(ws)
    if not rows: return _empty("Tes Lain", "Sheet kosong"), {}
    i_ujian = col(headers, "ujian"); i_ue1 = col(headers, "ue1")
    i_batch = col(headers, "batch"); i_tgl = col(headers, "tgl_ujian")
    i_hadir = col(headers, "kehadiran")
    i_tag   = col(headers, "tag")
    i_kat   = col(headers, "kategori")

    n_meng = n_tidak = n_belum = 0
    months_by_year = {}; rows_by_year = defaultdict(list)
    tag_ujian_by_year = {}  # {yr: {tag: {ujian: count}}}

    for r in rows:
        is_m = is_t = is_b = False
        if i_hadir != -1:
            h = val(r, i_hadir).lower().strip()
            if h == "mengikuti": n_meng += 1; is_m = True
            elif h == "tidak mengikuti": n_tidak += 1; is_t = True
            elif h == "belum diproses": n_belum += 1; is_b = True
        raw_tgl = r[i_tgl] if i_tgl != -1 and i_tgl < len(r) else None
        m, y = extract_month_year(raw_tgl)
        if y is None: y = datetime.now().year
        tag   = val(r, i_tag).strip().lower() if i_tag != -1 else ""
        ujian = val(r, i_ujian).strip()
        if is_m: add_month_year(months_by_year, raw_tgl)
        kategori = val(r, i_kat).strip() if i_kat != -1 else ""
        rows_by_year[y].append({
            "mengikuti": is_m, "tidak_mengikuti": is_t, "belum_diproses": is_b,
            "ujian": ujian, "ue1": val(r, i_ue1), "batch": val(r, i_batch),
            "tag": tag, "kategori": kategori, "month": m, "year": y,
        })
        # Agregat tag × ujian (hanya peserta yang mengikuti)
        if is_m and (tag or ujian):
            if y not in tag_ujian_by_year: tag_ujian_by_year[y] = {}
            t_key = tag if tag else "(tanpa tag)"
            u_key = ujian if ujian else "(tanpa ujian)"
            if t_key not in tag_ujian_by_year[y]: tag_ujian_by_year[y][t_key] = {}
            tag_ujian_by_year[y][t_key][u_key] = tag_ujian_by_year[y][t_key].get(u_key, 0) + 1

    total = len(rows)
    hby = {yr: sum(1 for r in rl if r["mengikuti"]) for yr, rl in rows_by_year.items()}
    fby = {}
    for yr, rl in rows_by_year.items():
        fby[str(yr)] = {
            "terdaftar": len(rl), "mengikuti": sum(1 for r in rl if r["mengikuti"]),
            "tidak_mengikuti": sum(1 for r in rl if r["tidak_mengikuti"]),
            "belum_diproses": sum(1 for r in rl if r["belum_diproses"]),
        }
    log.info(f"[Tes Lain] {total} terdaftar, {n_meng} mengikuti")
    return {
        "team": "Tes Lain", "type": "count", "total": total, "total_hadir": n_meng,
        "lulus": 0, "tidak_lulus": 0, "pct_lulus": 0,
        "total_angkatan": len(set(r["batch"] for rl in rows_by_year.values() for r in rl if r["mengikuti"] and r["batch"])),
        "targets": targets.get("Tes Lain", {}), "months_by_year": months_by_year,
        "hadir_by_year": hby, "funnel_by_year": fby,
        "tag_ujian_by_year": {str(yr): d for yr, d in tag_ujian_by_year.items()},
        "error": None,
    }, dict(rows_by_year)


def _parse_date_mdy(raw):
    """Parse tanggal format mm/dd/yyyy. Return (month, year) atau (None, None)."""
    if raw is None: return None, None
    # datetime/date object langsung
    if hasattr(raw, 'month'): return raw.month, raw.year
    # Excel serial number
    if isinstance(raw, (int, float)) and 30000 < raw < 60000:
        from datetime import date, timedelta
        serial = int(raw)
        if serial > 59: serial -= 1
        d = date(1899, 12, 31) + timedelta(days=serial)
        return d.month, d.year
    s = str(raw).strip()
    if not s or s.lower() == 'none': return None, None
    # Format mm/dd/yyyy
    for sep in ['/', '-', '.']:
        parts = s.split(sep)
        if len(parts) == 3:
            try:
                p = [int(x) for x in parts]
                # mm/dd/yyyy: p[2] adalah tahun 4 digit
                if p[2] > 1900:
                    return p[0], p[2]
                # yyyy/mm/dd
                if p[0] > 1900:
                    return p[1], p[0]
            except: pass
    return None, None


def _raw_cell(row, idx):
    """Ambil nilai mentah cell (bukan string) untuk parsing tanggal."""
    if idx == -1 or idx >= len(row): return None
    return row[idx]


def _has_date(row, idx):
    """True jika cell di idx punya nilai tanggal yang valid."""
    raw = _raw_cell(row, idx)
    if raw is None: return False
    m, y = _parse_date_mdy(raw)
    return m is not None


def _sub_exam_base(name, label_lulus, label_tidak):
    return {
        "name": name, "total": 0, "lulus": 0, "tidak_lulus": 0, "pct_lulus": 0,
        "label_lulus": label_lulus, "label_tidak": label_tidak,
        "col_tgl2_label": "Wawancara",   # default, di-override per sheet
        "by_periode": [], "error": None,
    }


def _finalize_sub(name, label_lulus, label_tidak, col_tgl2_label,
                  total, lulus_cnt, tidak_cnt, periode_map):
    """Hitung agregat akhir dan susun by_periode."""
    # pct dari yang sudah ada hasil saja (lulus+tidak_lulus), bukan dari total NIP
    ada_hasil = lulus_cnt + tidak_cnt
    pct = round(lulus_cnt / ada_hasil * 100, 1) if ada_hasil > 0 else 0
    by_periode = []
    for periode, d in sorted(periode_map.items()):
        p_ada = d["lulus"] + d["tidak_lulus"]
        p_pct = round(d["lulus"] / p_ada * 100, 1) if p_ada > 0 else 0
        by_periode.append({
            "periode":     periode,
            "total":       d["total"],
            "ukom_teknis": d["ukom_teknis"],
            "tgl2":        d["tgl2"],
            "mansoskul":   d["mansoskul"],
            "lulus":       d["lulus"],
            "tidak_lulus": d["tidak_lulus"],
            "pct_lulus":   p_pct,
        })
    log.info(f"[{name}] {total} peserta (NIP), {lulus_cnt} {label_lulus}, "
             f"{tidak_cnt} {label_tidak} ({pct}%), {len(by_periode)} periode")
    return {
        "name": name, "total": total, "lulus": lulus_cnt,
        "tidak_lulus": tidak_cnt, "pct_lulus": pct,
        "label_lulus": label_lulus, "label_tidak": label_tidak,
        "col_tgl2_label": col_tgl2_label,
        "by_periode": by_periode, "error": None,
    }


def _make_periode_entry():
    return {"total": 0, "ukom_teknis": 0, "tgl2": 0, "mansoskul": 0,
            "lulus": 0, "tidak_lulus": 0}


def _process_sub_generic(ws, name, lbl_l, lbl_t, col_hasil_name,
                          col_tgl1, col_tgl2, col_tgl3, col_tgl2_label,
                          is_lulus_fn):
    """
    Processor generik sub-exam:
    - total peserta = jumlah baris yang punya NIP (semua baris termasuk yang belum ada hasil)
    - lulus/tidak_lulus = dari kolom hasil (baris tanpa hasil tidak dihitung ke lulus/tidak)
    - by_periode = agregat per kolom Periode (kosong → '-')
    """
    headers, rows = to_rows(ws)
    if not rows:
        return {**_sub_exam_base(name, lbl_l, lbl_t), "error": "Sheet kosong"}

    i_hasil   = col(headers, col_hasil_name)
    i_nip     = col(headers, "nip")
    i_periode = col(headers, "periode")
    i_tgl1    = col(headers, col_tgl1)
    i_tgl2    = col(headers, col_tgl2)
    i_tgl3    = col(headers, col_tgl3)

    if i_hasil == -1:
        return {**_sub_exam_base(name, lbl_l, lbl_t),
                "error": f"Kolom '{col_hasil_name}' tidak ditemukan"}

    total = lulus_cnt = tidak_cnt = 0
    periode_map = {}

    for r in rows:
        # Total peserta: hitung dari NIP. Jika kolom NIP tidak ada, hitung semua baris.
        nip_val = val(r, i_nip).strip() if i_nip != -1 else ""
        if i_nip != -1 and not nip_val:
            continue   # baris tanpa NIP = bukan peserta, lewati
        total += 1

        v = val(r, i_hasil).strip()
        is_lul = is_lulus_fn(v) if v else None  # None = belum ada hasil

        if is_lul is True:  lulus_cnt += 1
        elif is_lul is False: tidak_cnt += 1
        # is_lul None → belum ada hasil, tidak masuk lulus/tidak_lulus

        periode = val(r, i_periode).strip() if i_periode != -1 else ""
        if not periode: periode = "-"
        if periode not in periode_map: periode_map[periode] = _make_periode_entry()
        d = periode_map[periode]
        d["total"] += 1
        if _has_date(r, i_tgl1): d["ukom_teknis"] += 1
        if _has_date(r, i_tgl2): d["tgl2"] += 1
        if _has_date(r, i_tgl3): d["mansoskul"] += 1
        if is_lul is True:  d["lulus"] += 1
        elif is_lul is False: d["tidak_lulus"] += 1

    return _finalize_sub(name, lbl_l, lbl_t, col_tgl2_label,
                         total, lulus_cnt, tidak_cnt, periode_map)


def process_sub_PKAPK(ws):
    """PKAPK APBN: Hasil UKOM → 'Lulus'. Tanggal: ukom_teknis, wawancara, mansoskul."""
    return _process_sub_generic(
        ws, "PKAPK APBN", "Lulus", "Tidak Lulus",
        "hasil ukom", "ukom_teknis", "wawancara", "mansoskul", "Wawancara",
        lambda v: v.lower() == "lulus",
    )


def process_sub_Pelelang(ws):
    """Pelelang: Hasil UKOM Teknis → 'Lulus'. Tanggal: ukom_teknis, wawancara, mansoskul."""
    return _process_sub_generic(
        ws, "Pelelang", "Lulus", "Tidak Lulus",
        "hasil ukom teknis", "ukom_teknis", "wawancara", "mansoskul", "Wawancara",
        lambda v: v.lower() == "lulus",
    )


def process_sub_AKPD(ws):
    """AKPD: Hasil UKOM Teknis → 'Kompeten'/'Belum Kompeten'. Tanggal: ukom_teknis, ukom_tertulis, mansoskul."""
    return _process_sub_generic(
        ws, "AKPD", "Kompeten", "Belum Kompeten",
        "hasil ukom teknis", "ukom_teknis", "ukom_tertulis", "mansoskul", "UKOM Tertulis",
        lambda v: v.lower() == "kompeten" and "belum" not in v.lower(),
    )


def process_sub_AA(ws):
    """AA: Hasil → 'LULUS' (case-insensitive). Tanggal: ukom_teknis, wawancara, mansoskul."""
    return _process_sub_generic(
        ws, "AA", "Lulus", "Tidak Lulus",
        "hasil", "ukom_teknis", "wawancara", "mansoskul", "Wawancara",
        lambda v: v.lower() == "lulus",
    )


SUB_EXAM_PROCESSORS = {
    "PKAPK APBN": process_sub_PKAPK,
    "Pelelang":   process_sub_Pelelang,
    "AKPD":       process_sub_AKPD,
    "AA":         process_sub_AA,
}



PROCESSORS = {
    "JFKN": process_JFKN, "SAK": process_SAK, "AC": process_AC,
    "Beasiswa": process_Beasiswa, "USKP": process_USKP, "PBJ": process_PBJ,
    "Tes Lain": process_TesTes,
}


# =============================================================
#  PROSES SEMUA SHEET — output 2-layer
# =============================================================


def process_Akreditasi(ws):
    """Baca sheet Akreditasi — hasilkan aggregat + daftar program."""
    headers, rows = to_rows(ws)
    if not rows:
        return {"lembaga": 0, "program": 0, "dalam_proses": 0, "items": []}

    i_lembaga = col(headers, "lembaga penyelenggara pelatihan pemerintah", "lembaga")
    i_program  = col(headers, "program pelatihan teknis", "program")
    i_status   = col(headers, "status")
    i_ket      = col(headers, "keterangan")

    items = []
    lembaga_set = set()
    lembaga_ter_set = set()
    n_terakreditasi = 0
    n_proses = 0

    for r in rows:
        lembaga = val(r, i_lembaga).strip()
        program = val(r, i_program).strip()
        status  = val(r, i_status).strip()
        ket     = val(r, i_ket).strip()
        if not lembaga and not program:
            continue
        lembaga_set.add(lembaga)
        status_lower = status.lower()
        is_ter = "terakreditasi" in status_lower and "tahap" not in status_lower
        is_pros = "tahap" in status_lower or "proses" in status_lower
        if is_ter:
            n_terakreditasi += 1
            lembaga_ter_set.add(lembaga)
        if is_pros:
            n_proses += 1
        items.append({
            "lembaga": lembaga, "program": program,
            "status": status, "keterangan": ket,
            "is_terakreditasi": is_ter, "is_proses": is_pros,
        })

    return {
        "lembaga": len(lembaga_ter_set),
        "program": n_terakreditasi,
        "dalam_proses": n_proses,
        "items": items,
    }

def process_all_sheets():
    if not os.path.exists(EXCEL_FILE):
        log.error(f"File tidak ditemukan: {EXCEL_FILE}"); return
    try:
        log.info(f"Membuka: {os.path.basename(EXCEL_FILE)}")
        wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
        log.info(f"Sheet tersedia: {wb.sheetnames}")
        targets = read_targets(wb)
        teams_summary = []; detail_files = {}

        for team_name, sheet_name in SHEET_MAP.items():
            if sheet_name not in wb.sheetnames:
                log.warning(f"[{team_name}] Sheet '{sheet_name}' tidak ditemukan.")
                teams_summary.append(_empty(team_name, f"Sheet '{sheet_name}' tidak ditemukan"))
                continue
            proc = PROCESSORS.get(team_name)
            if not proc: teams_summary.append(_empty(team_name)); continue
            summary, rby = proc(wb[sheet_name], targets)
            teams_summary.append(summary)
            slug = safe_slug(team_name)
            for yr, yr_rows in rby.items():
                detail_files[f"detail_{slug}_{yr}.json"] = yr_rows
        # Sub-exam JFKN (PKAPK APBN, Pelelang, AKPD, AA)
        jfkn_sub_exams = {}
        for sub_name, proc_fn in SUB_EXAM_PROCESSORS.items():
            sheet_n = JFKN_SUB_SHEETS[sub_name]["sheet"]
            if sheet_n not in wb.sheetnames:
                log.warning(f"[JFKN sub] Sheet '{sheet_n}' tidak ditemukan — dilewati")
                cfg_s = JFKN_SUB_SHEETS[sub_name]
                jfkn_sub_exams[sub_name] = {
                    "name": sub_name, "total": 0, "lulus": 0, "tidak_lulus": 0,
                    "pct_lulus": 0, "label_lulus": cfg_s["label_lulus"],
                    "label_tidak": cfg_s["label_tidak"],
                    "col_tgl2_label": "Wawancara", "by_periode": [],
                    "error": f"Sheet '{sheet_n}' tidak ditemukan",
                }
            else:
                jfkn_sub_exams[sub_name] = proc_fn(wb[sheet_n])
        # Tempelkan ke summary JFKN
        for t in teams_summary:
            if t.get("team") == "JFKN":
                t["sub_exams"] = jfkn_sub_exams
                break

        # Akreditasi — sheet opsional, tidak masuk ke teams
        akreditasi = {}
        if "Akreditasi" in wb.sheetnames:
            akreditasi = process_Akreditasi(wb["Akreditasi"])
            log.info(f"[Akreditasi] {akreditasi['lembaga']} lembaga, {akreditasi['program']} program terakreditasi")
        else:
            log.warning("Sheet 'Akreditasi' tidak ditemukan — dilewati")

        wb.close()
    except Exception as e:
        log.error(f"Gagal membuka file Excel: {e}"); import traceback; traceback.print_exc(); return

    cur_year = datetime.now().year

    # Agregasi monthly_by_year dari semua tim, exclude Beasiswa
    EXCLUDE_MONTHLY = {"Beasiswa"}
    monthly_agg = {}
    for t in teams_summary:
        if t.get("team") in EXCLUDE_MONTHLY:
            continue
        for yr, months in t.get("months_by_year", {}).items():
            yr_key = str(yr)
            if yr_key not in monthly_agg:
                monthly_agg[yr_key] = {}
            for m, mdata in months.items():
                m_key = str(m)
                monthly_agg[yr_key][m_key] = monthly_agg[yr_key].get(m_key, 0) + mdata.get("total", 0)
    # Urutkan bulan per tahun
    monthly_agg = {yr: dict(sorted(ms.items(), key=lambda x: int(x[0]))) for yr, ms in sorted(monthly_agg.items())}

    pct_teams = [t for t in teams_summary if t.get("type") == "pct"]
    total_peserta = sum(t.get("total", 0) for t in pct_teams)
    total_lulus = sum(t.get("lulus", 0) for t in pct_teams)
    total_hadir = sum(t.get("total_hadir", t.get("total", 0)) for t in teams_summary)
    total_target = sum(t.get("targets", {}).get(cur_year, 0) for t in teams_summary)
    beasiswa_t = next((t for t in teams_summary if t["team"] == "Beasiswa"), None)
    ac_t = next((t for t in teams_summary if t["team"] == "AC"), None)

    output = {
        "generated_at": datetime.now().strftime("%d %B %Y, %H:%M:%S"),
        "generated_ts": datetime.now().isoformat(),
        "total_peserta": total_peserta, "total_lulus": total_lulus,
        "avg_pct_lulus": round(total_lulus/total_peserta*100, 1) if total_peserta > 0 else 0,
        "total_tim": len(SHEET_MAP),
        "total_angkatan": sum(t.get("total_angkatan", 0) for t in teams_summary),
        "total_hadir": total_hadir, "total_target": total_target,
        "pct_capaian": round(total_hadir/total_target*100, 1) if total_target > 0 else 0,
        "tim_mencapai": sum(1 for t in teams_summary if t.get("targets",{}).get(cur_year,0) > 0 and t.get("total_hadir", t.get("total",0)) >= t["targets"][cur_year]),
        "total_beasiswa": beasiswa_t["total"] if beasiswa_t else 0,
        "total_beasiswa_selesai": beasiswa_t.get("total_selesai", 0) if beasiswa_t else 0,
        "total_ac": ac_t["total"] if ac_t else 0,
        "akreditasi": akreditasi,
        "monthly_by_year": monthly_agg,
        "teams": teams_summary,
    }

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    sp = os.path.join(OUTPUT_DIR, "summary.json")
    with open(sp, "w", encoding="utf-8") as f: json.dump(output, f, ensure_ascii=False, indent=2)
    log.info(f"summary.json -> {sp} ({os.path.getsize(sp)//1024} KB)")

    for fname, rl in detail_files.items():
        fp = os.path.join(OUTPUT_DIR, fname)
        with open(fp, "w", encoding="utf-8") as f: json.dump(rl, f, ensure_ascii=False)
        log.info(f"  {fname} -> {len(rl)} rows")
    log.info(f"Selesai: 1 summary + {len(detail_files)} detail files")


# =============================================================
#  FILE WATCHER
# =============================================================

class ExcelChangeHandler(FileSystemEventHandler):
    def __init__(self): self._last = 0
    def on_modified(self, event): self._handle(event)
    def on_created(self, event): self._handle(event)
    def _handle(self, event):
        if event.is_directory: return
        if os.path.normcase(os.path.abspath(event.src_path)) != os.path.normcase(os.path.abspath(EXCEL_FILE)): return
        now = time.time()
        if now - self._last < DEBOUNCE_SECONDS: return
        self._last = now
        log.info("Perubahan terdeteksi -> memproses ulang...")
        time.sleep(1.5)
        process_all_sheets()


def main():
    log.info("=" * 60)
    log.info("  Dashboard Monitoring Terpadu v2 — Watcher Aktif")
    log.info("  Arsitektur: summary.json + detail_{tim}_{thn}.json")
    log.info("=" * 60)
    log.info(f"  File  : {EXCEL_FILE}")
    log.info(f"  Output: {OUTPUT_DIR}")
    log.info(f"  Sheet : {list(SHEET_MAP.values())}")
    log.info("  Ctrl+C untuk berhenti.")
    log.info("=" * 60)
    if not os.path.exists(EXCEL_FILE):
        log.error(f"File tidak ditemukan: {EXCEL_FILE}"); sys.exit(1)
    log.info("Memproses data awal...")
    process_all_sheets()
    watch_dir = os.path.dirname(os.path.abspath(EXCEL_FILE))
    handler = ExcelChangeHandler()
    observer = Observer()
    observer.schedule(handler, path=watch_dir, recursive=False)
    observer.start()
    log.info(f"Watcher aktif — memantau: {watch_dir}")
    try:
        while True: time.sleep(1)
    except KeyboardInterrupt:
        log.info("Watcher dihentikan."); observer.stop()
    observer.join()

if __name__ == "__main__":
    main()